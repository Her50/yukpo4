#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Génère le référentiel Excel des fournitures (cahiers + accessoires) par classe.

Source : 31 listes PDF officielles du Complexe Scolaire Bilingue Soft Education
(année scolaire 2025-2026). Extraction faite par Claude le 2026-05-22.

Couvre :
  - Primaire francophone ordinaire : SIL, CP, CE1, CE2, CM1, CM2
  - Primaire francophone bilingue (5 nouvelles classes) : SIL BIL, CP BIL,
    CE1 BIL, CE2 BIL, CM1 BIL
  - Primaire anglophone : Class 1 à Class 6
  - Secondaire francophone 1er cycle : 6ème, 5ème, 4ème, 3ème
  - Secondaire francophone 2nd cycle : 2nde A-C, 2nde STT, 1ère
  - Secondaire anglophone : Form 1-5, Form 3/4 Commercial, Form 4 General,
    Lower Sixth, Upper Sixth

Les colonnes Prix Standard et Prix Haut de Gamme sont VIDES — à remplir
manuellement par l'utilisateur puis ré-importer pour seeder la table
accessoires_populaires_par_classe.

Convention : exclut les fournitures non-didactiques (sacs, EPI, tissus
couture, savon parfumé, etc.) — cf. memory feedback_seed_accessoires_didactique_only.
"""

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from pathlib import Path

# ─────────────────────────────────────────────────────────────────────────────
# DONNÉES — extraites des 31 PDFs
# Format : { classe: [ (categorie, article, quantite, specification), ... ] }
# ─────────────────────────────────────────────────────────────────────────────

# Catégories : "cahier" | "ecriture" | "geometrie" | "ardoise" | "protection"
#              | "dictionnaire" | "papier" | "art" | "rangement"

DATA = {
    # ────────── PRIMAIRE FRANCOPHONE ORDINAIRE ──────────
    "SIL": [
        ("cahier", "Paquet cahiers bananier 24 cahiers", 1, "doubles lignes"),
        ("cahier", "Cahier dessin petit format", 1, ""),
        ("cahier", "Cahier 60 leaves plain lines (English)", 2, ""),
        ("cahier", "Cahier de texte pour devoir à domicile", 1, ""),
        ("ecriture", "Paquet crayons ordinaires 2B ou HB", 2, ""),
        ("ecriture", "Taille-crayon", 2, ""),
        ("ecriture", "Gomme", 10, "blanche"),
        ("ecriture", "Tube de colle UHU", 1, ""),
        ("ecriture", "Paquet marqueurs Bic Velleda pour ardoise", 1, ""),
        ("ecriture", "Paquet crayons couleur", 1, ""),
        ("ardoise", "Ardoise à feutre", 1, ""),
        ("ardoise", "Effaçoir pour ardoise", 2, ""),
        ("geometrie", "Règle graduée 30cm", 1, ""),
        ("rangement", "Trousse", 1, ""),
        ("rangement", "Chemise à rabat", 1, ""),
        ("protection", "Couverture pour cahier", 24, "6 bleues, 6 rouges, 6 jaunes, 6 vertes"),
        ("papier", "Paquet papier canson", 1, ""),
    ],
    "CP": [
        ("cahier", "Cahier ordinaire 200 pages", 2, ""),
        ("cahier", "Cahier 144 pages", 6, ""),
        ("cahier", "Cahier 80 leaves plain lines (English)", 3, ""),
        ("cahier", "Cahier 40 leaves red and blue lines", 1, ""),
        ("cahier", "Cahier doubles lignes bananier", 15, ""),
        ("cahier", "Cahier dessin petit format", 1, ""),
        ("cahier", "Cahier de texte", 1, "pour devoir à domicile"),
        ("cahier", "Cahier de liaison", 1, ""),
        ("ecriture", "Stylo bleu", 6, ""),
        ("ecriture", "Stylo rouge", 4, ""),
        ("ecriture", "Paquet crayons HB ou 2B", 1, ""),
        ("ecriture", "Paquet crayons couleur", 1, ""),
        ("ecriture", "Taille-crayon", 1, ""),
        ("ecriture", "Gomme", 5, ""),
        ("ecriture", "Colle UHU", 1, ""),
        ("ecriture", "Paquet marqueurs Bic Velleda + crayons couleur", 1, ""),
        ("ardoise", "Ardoise à feutre", 1, ""),
        ("ardoise", "Effaçoir pour ardoise", 1, ""),
        ("geometrie", "Règle graduée 30cm", 1, ""),
        ("rangement", "Trousse", 1, ""),
        ("rangement", "Chemise à rabat", 1, ""),
        ("protection", "Couverture pour cahier", 20, "5 bleues, 5 rouges, 5 jaunes, 5 vertes"),
        ("protection", "Couvre-livre", 1, "paquet"),
    ],
    "CE1": [
        ("cahier", "Cahier français 200 pages", 1, ""),
        ("cahier", "Cahier d'exercices français 200 pages", 1, ""),
        ("cahier", "Cahier de devoirs 200 pages", 1, ""),
        ("cahier", "Cahier mathématiques 288 pages", 1, ""),
        ("cahier", "Cahier d'exercices maths 200 pages", 1, ""),
        ("cahier", "Cahier d'informatique 200 pages", 1, ""),
        ("cahier", "Cahier de chants et récits 100 pages", 1, ""),
        ("cahier", "Cahier de résumés 200 pages", 1, ""),
        ("cahier", "Cahier de contrôle hebdomadaire 288 pages", 1, ""),
        ("cahier", "Cahier de texte", 1, ""),
        ("cahier", "Cahier de liaison", 1, ""),
        ("cahier", "Cahier 40 leaves plain lines English", 2, ""),
        ("cahier", "Cahier de dessin grand format", 1, ""),
        ("ecriture", "Stylo bleu", 2, ""),
        ("ecriture", "Stylo rouge", 2, ""),
        ("ecriture", "Crayon ordinaire 2B ou HB", 2, ""),
        ("ecriture", "Taille-crayon", 1, ""),
        ("ecriture", "Gomme", 1, ""),
        ("ecriture", "Clé USB", 1, ""),
        ("ecriture", "Paquet feutres + crayons couleur", 1, ""),
        ("ardoise", "Ardoise à feutre", 1, ""),
        ("ardoise", "Paquet feutres pour ardoise", 1, ""),
        ("ardoise", "Effaçoir", 1, ""),
        ("geometrie", "Règle graduée 30cm", 1, ""),
        ("geometrie", "Boîte académique", 1, ""),
        ("rangement", "Trousse", 1, ""),
        ("rangement", "Chemise à rabat", 1, ""),
        ("protection", "Couverture pour cahier", 12, "3 vertes, 3 rouges, 3 jaunes, 3 bleues"),
    ],
    "CE2": [
        ("cahier", "Cahier français 200 pages", 1, ""),
        ("cahier", "Cahier d'exercices français 200 pages", 1, ""),
        ("cahier", "Cahier de devoirs 288 pages", 1, ""),
        ("cahier", "Cahier mathématiques 288 pages", 1, ""),
        ("cahier", "Cahier d'exercices maths 200 pages", 1, ""),
        ("cahier", "Cahier d'informatique 144 pages", 1, ""),
        ("cahier", "Cahier de chants et récits 100 pages", 1, ""),
        ("cahier", "Cahier de résumés 288 pages", 1, ""),
        ("cahier", "Cahier de contrôle hebdomadaire 288 pages", 1, ""),
        ("cahier", "Cahier de dictée et PE 288 pages", 1, ""),
        ("cahier", "Cahier 40 leaves plain lines English", 2, ""),
        ("cahier", "Cahier de texte", 1, ""),
        ("cahier", "Cahier de dessin grand format", 1, ""),
        ("ecriture", "Stylo bleu", 2, ""),
        ("ecriture", "Stylo rouge", 2, ""),
        ("ecriture", "Crayon ordinaire 2B ou HB", 2, ""),
        ("ecriture", "Taille-crayon", 1, ""),
        ("ecriture", "Paquet crayons couleur + feutres", 1, ""),
        ("ardoise", "Ardoise à feutre", 1, ""),
        ("ardoise", "Paquet feutres pour ardoise", 1, ""),
        ("ardoise", "Effaçoir", 1, ""),
        ("geometrie", "Règle graduée 30cm", 1, ""),
        ("geometrie", "Boîte académique", 1, ""),
        ("dictionnaire", "Dictionnaire français", 1, ""),
        ("rangement", "Trousse", 1, ""),
        ("rangement", "Chemise à rabat", 1, ""),
        ("protection", "Couverture pour cahier", 12, "3 vertes, 3 rouges, 3 jaunes, 3 bleues"),
    ],
    "CM1": [
        ("cahier", "Cahier français 288 pages", 1, "couverture verte"),
        ("cahier", "Cahier d'exercices français 300 pages", 1, "couverture verte"),
        ("cahier", "Cahier mathématiques 288 pages", 1, "couverture bleue"),
        ("cahier", "Cahier d'exercices maths 300 pages", 1, "couverture bleue"),
        ("cahier", "Cahier de résumés 300 pages", 1, "couverture marron"),
        ("cahier", "Cahier d'informatique 100 pages", 1, "couverture rouge"),
        ("cahier", "Cahier de chants/récits 100 pages", 1, "couverture rouge"),
        ("cahier", "Cahier 80 leaves plain lines English", 2, "couverture rose"),
        ("cahier", "Cahier de dictée et PE 288 pages", 1, "couverture verte"),
        ("cahier", "Cahier de texte", 1, ""),
        ("cahier", "Cahier de liaison", 1, ""),
        ("cahier", "Cahier d'évaluation hebdomadaire 300 pages", 1, "couverture orange"),
        ("cahier", "Cahier TP Sciences/Technologie 200 pages", 1, "couverture verte"),
        ("cahier", "Cahier de dessin grand format", 1, ""),
        ("ecriture", "Stylo bleu", 2, ""),
        ("ecriture", "Stylo rouge", 2, ""),
        ("ecriture", "Paquet crayons ordinaires 2B", 1, ""),
        ("ecriture", "Taille-crayon", 1, ""),
        ("ecriture", "Gomme", 1, ""),
        ("ecriture", "Bâton de colle", 1, ""),
        ("ecriture", "Clé USB", 1, ""),
        ("ecriture", "Paquet marqueurs ardoise + crayons couleur", 1, ""),
        ("ardoise", "Ardoise à marqueurs", 1, ""),
        ("ardoise", "Effaçoir", 1, ""),
        ("geometrie", "Règle graduée 30cm", 1, ""),
        ("geometrie", "Boîte académique", 1, ""),
        ("dictionnaire", "Dictionnaire français", 1, ""),
        ("rangement", "Trousse", 1, ""),
        ("rangement", "Chemise à rabat", 1, ""),
    ],
    "CM2": [
        ("cahier", "Cahier français 288 pages", 1, "couverture verte"),
        ("cahier", "Cahier d'exercices français 288 pages", 1, "couverture verte"),
        ("cahier", "Cahier d'exercices maths 288 pages", 1, "couverture verte"),
        ("cahier", "Cahier mathématiques 288 pages", 1, "couverture bleue"),
        ("cahier", "Cahier de résumés 300 pages", 1, "couverture marron"),
        ("cahier", "Cahier de dessin A4 grand format", 1, ""),
        ("cahier", "Cahier d'informatique 144 pages", 1, "couverture rouge"),
        ("cahier", "Cahier de chants et récits 100 pages", 1, "couverture rouge"),
        ("cahier", "Cahier de dictée et PE 288 pages", 1, "couverture verte"),
        ("cahier", "Cahier de texte pour devoir", 1, ""),
        ("cahier", "Cahier de liaison", 1, ""),
        ("cahier", "Cahier d'évaluation hebdomadaire 300 pages", 1, "couverture orange"),
        ("cahier", "Cahier de recherche 200 pages", 1, "couverture verte"),
        ("cahier", "Cahier TP Sciences/Tech 200 pages", 1, "couverture verte"),
        ("cahier", "Cahier 60 leaves plain lines English", 2, ""),
        ("ecriture", "Stylo bleu", 2, ""),
        ("ecriture", "Stylo rouge", 2, ""),
        ("ecriture", "Crayon ordinaire 2B", 2, ""),
        ("ecriture", "Taille-crayon", 1, ""),
        ("ecriture", "Gomme", 1, ""),
        ("ecriture", "Bâton de colle", 1, ""),
        ("ecriture", "Clé USB", 1, ""),
        ("ecriture", "Paquet marqueurs ardoise + crayons couleur", 1, ""),
        ("ardoise", "Ardoise à marqueurs", 1, ""),
        ("ardoise", "Effaçoir", 1, ""),
        ("geometrie", "Règle graduée 30cm", 1, ""),
        ("geometrie", "Boîte académique", 1, ""),
        ("dictionnaire", "Dictionnaire français", 1, ""),
        ("rangement", "Trousse", 1, ""),
        ("rangement", "Chemise à rabat", 1, ""),
    ],
    # ────────── PRIMAIRE BILINGUE (5 nouvelles classes) ──────────
    "SIL BIL": [
        ("cahier", "Cahier dessin petit format", 1, ""),
        ("cahier", "Paquet cahiers bananier 24 cahiers", 1, ""),
        ("cahier", "Cahier 80 leaves plain lines (Science/Tech)", 3, ""),
        ("cahier", "Cahier 80 leaves square lines (Maths)", 2, ""),
        ("cahier", "Cahier 80 leaves plain lines (English/SWB)", 3, ""),
        ("cahier", "Cahier 60 leaves plain lines (S/S)", 2, ""),
        ("cahier", "Cahier 80 leaves plain lines (V/S)", 1, ""),
        ("cahier", "Cahier red and blue lines (Writing)", 1, ""),
        ("cahier", "Cahier 40 leaves plain lines (Sports/Music)", 2, ""),
        ("cahier", "Cahier 60 leaves plain lines (ICT/NLC)", 2, ""),
        ("cahier", "Cahier 80 leaves plain lines (Homework)", 2, ""),
        ("cahier", "Cahier de texte for homework", 1, ""),
        ("cahier", "Drawing book A4", 1, ""),
        ("ecriture", "Paquet crayons (Bic Evolution)", 2, ""),
        ("ecriture", "Taille-crayon", 5, ""),
        ("ecriture", "Gomme", 10, ""),
        ("ecriture", "Tube colle UHU + UHU glue stick", 1, ""),
        ("ecriture", "Paquets marqueurs Bic Velleda", 2, ""),
        ("ecriture", "Paquet crayons couleur", 2, ""),
        ("ecriture", "Paquet patafix", 1, ""),
        ("ardoise", "Ardoise à feutre", 1, ""),
        ("ardoise", "Effaçoir / Duster", 1, ""),
        ("geometrie", "Règle graduée 30cm", 2, ""),
        ("rangement", "Trousse", 1, ""),
        ("rangement", "Chemise à rabat", 1, ""),
        ("protection", "Couverture pour cahier", 24, "6 bleues, 6 rouges, 6 jaunes, 6 vertes"),
        ("protection", "Couvre-livre transparent", 20, ""),
        ("papier", "Paquet papier canson", 1, ""),
    ],
    "CP BIL": [
        ("cahier", "Cahier 144 pages", 6, ""),
        ("cahier", "Cahier 200 pages", 2, ""),
        ("cahier", "Cahier dessin petit format", 1, ""),
        ("cahier", "Cahier 80 leaves plain lines (English/SWB)", 3, ""),
        ("cahier", "Cahier 80 leaves square lines (Maths)", 2, ""),
        ("cahier", "Cahier 80 leaves plain lines (S/T)", 2, ""),
        ("cahier", "Cahier 60 leaves plain lines (S/S)", 2, ""),
        ("cahier", "Cahier 80 leaves plain lines (V/S)", 1, ""),
        ("cahier", "Cahier 60 leaves plain lines (ICT/NLC)", 2, ""),
        ("cahier", "Cahier 40 leaves plain lines (Sport/Music)", 2, ""),
        ("cahier", "Cahier 20 leaves red and blue lines", 1, ""),
        ("cahier", "Cahier doubles lignes bananier", 15, ""),
        ("cahier", "Cahier de texte", 1, ""),
        ("cahier", "Cahier de liaison", 1, ""),
        ("ecriture", "Stylo bleu", 6, ""),
        ("ecriture", "Stylo noir", 2, ""),
        ("ecriture", "Stylo rouge", 2, ""),
        ("ecriture", "Taille-crayon", 5, ""),
        ("ecriture", "Gomme", 5, ""),
        ("ecriture", "Tube colle UHU + UHU glue stick", 1, ""),
        ("ecriture", "Paquets marqueurs Bic Velleda", 2, ""),
        ("ecriture", "Paquet crayons couleur", 1, ""),
        ("ecriture", "Paquet crayons 2B ou HB", 1, ""),
        ("ecriture", "Paquet patafix", 1, ""),
        ("ardoise", "Ardoise à feutre", 1, ""),
        ("ardoise", "Effaçoir / Duster", 1, ""),
        ("geometrie", "Règle graduée 30cm", 2, ""),
        ("rangement", "Trousse", 1, ""),
        ("rangement", "Chemise à rabat", 1, ""),
        ("protection", "Couverture pour cahier", 20, "5 bleues, 5 rouges, 5 jaunes, 5 vertes"),
        ("protection", "Couvre-livre transparent", 20, ""),
        ("papier", "Paquet papier canson (big size)", 1, ""),
    ],
    "CE1 BIL": [
        ("cahier", "Cahier français 200 pages", 1, ""),
        ("cahier", "Cahier d'exercices français 200 pages", 1, ""),
        ("cahier", "Cahier de devoirs 200 pages", 1, ""),
        ("cahier", "Cahier mathématiques 288 pages", 1, ""),
        ("cahier", "Cahier d'exercices maths 200 pages", 1, ""),
        ("cahier", "Cahier d'informatique 200 pages", 1, ""),
        ("cahier", "Cahier de chants et récits 120 pages", 1, ""),
        ("cahier", "Cahier de résumés 200 pages", 1, ""),
        ("cahier", "Cahier 80 leaves square lines (Mathematics)", 2, ""),
        ("cahier", "Cahier 80 leaves plain lines (English/Evaluation)", 3, ""),
        ("cahier", "Cahier 60 leaves plain lines (matières mineures)", 10, "S/W building, Health, Environmental, Tech, ICT, History, Geography, Peace, Home Ec"),
        ("cahier", "Cahier 40 leaves plain lines (matières mineures)", 7, "Civics, NLC, Human R, Agro-Past, Arts, Moral, Music"),
        ("cahier", "Cahier red and blue lines", 2, ""),
        ("cahier", "Cahier de dessin grand format / Drawing book A4", 1, ""),
        ("cahier", "Cahier de texte pour devoir", 1, ""),
        ("ecriture", "Stylo bleu", 2, ""),
        ("ecriture", "Stylo rouge", 2, ""),
        ("ecriture", "Paquet crayons 2B ou HB", 1, ""),
        ("ecriture", "Taille-crayon", 1, ""),
        ("ecriture", "Gomme", 1, ""),
        ("ecriture", "UHU glue stick and tube", 1, ""),
        ("ecriture", "Paquet feutres + crayons couleur", 1, ""),
        ("ecriture", "Paquet bold markers", 1, ""),
        ("ardoise", "Ardoise à feutre", 1, ""),
        ("ardoise", "Effaçoir", 1, ""),
        ("ardoise", "White tablet", 1, ""),
        ("geometrie", "Règle graduée 30cm", 1, ""),
        ("geometrie", "Boîte académique", 1, ""),
        ("rangement", "Trousse", 1, ""),
        ("rangement", "Chemise à rabat", 1, ""),
        ("protection", "Couverture pour cahier", 12, "3 vertes, 3 rouges, 3 jaunes, 3 bleues"),
    ],
    "CE2 BIL": [
        ("cahier", "Cahier français 200 pages", 2, ""),
        ("cahier", "Cahier d'exercices français 200 pages", 1, ""),
        ("cahier", "Cahier de devoirs 288 pages", 1, ""),
        ("cahier", "Cahier mathématiques 244 pages", 1, ""),
        ("cahier", "Cahier d'exercices maths 200 pages", 1, ""),
        ("cahier", "Cahier d'informatique 200 pages", 1, ""),
        ("cahier", "Cahier de chants et récits 120 pages", 1, ""),
        ("cahier", "Cahier de résumés 288 pages", 1, ""),
        ("cahier", "Cahier 80 leaves square lines (Maths)", 2, ""),
        ("cahier", "Cahier 80 leaves plain lines (English, History, Geography, Eval, SWB)", 6, ""),
        ("cahier", "Cahier 60 leaves plain lines (matières mineures)", 8, "Health Ed, Environmental, ICT, Home Ec, Tech, Civics, Agro-Past, Arts"),
        ("cahier", "Cahier 40 leaves plain lines", 5, "Human Right, Moral Ed, NLC, Sports, Agro-Past"),
        ("cahier", "Cahier de dessin grand format", 1, ""),
        ("cahier", "Cahier de texte", 1, ""),
        ("cahier", "Cahier de liaison", 1, ""),
        ("ecriture", "Stylo bleu", 2, ""),
        ("ecriture", "Stylo rouge", 2, ""),
        ("ecriture", "Paquet crayons 2B ou HB", 1, ""),
        ("ecriture", "Taille-crayon", 1, ""),
        ("ecriture", "UHU glue stick", 1, ""),
        ("ecriture", "Tube colle UHU", 1, ""),
        ("ecriture", "Paquet crayons couleur", 1, ""),
        ("ecriture", "Paquet marqueurs ardoise", 1, ""),
        ("ardoise", "Ardoise à feutre", 1, ""),
        ("ardoise", "Effaçoir", 1, ""),
        ("geometrie", "Règle graduée 30cm", 1, ""),
        ("geometrie", "Boîte académique", 1, ""),
        ("dictionnaire", "Dictionnaire français", 1, ""),
        ("rangement", "Trousse", 1, ""),
        ("rangement", "Chemise à rabat", 1, ""),
        ("protection", "Couverture pour cahier", 12, "3 vertes, 3 rouges, 3 jaunes, 3 bleues"),
    ],
    "CM1 BIL": [
        ("cahier", "Cahier 200 leaves ledger plain lines (English)", 2, ""),
        ("cahier", "Cahier 200 leaves ledger square lines (Maths)", 2, ""),
        ("cahier", "Cahier 80 leaves plain lines (matières majeures)", 10, "History, Geography, Health, Dictation, Eval, Tech/Eng, ICT, Home Ec"),
        ("cahier", "Cahier 60 leaves plain lines (matières mineures)", 6, "Civics, Agro Pastoral, Arts & Craft, SWB, Music, Sports"),
        ("cahier", "Cahier 40 leaves plain lines", 5, "Human Right, Moral Ed, visual Arts, NLC"),
        ("cahier", "Cahier de français 144 pages", 2, ""),
        ("cahier", "Drawing book A4", 1, ""),
        ("rangement", "Pencil case", 1, ""),
        ("protection", "Couvre-livre transparent", 20, "rouge, bleu, jaune, vert"),
    ],
    # ────────── PRIMAIRE ANGLOPHONE (CLASS 1-6) ──────────
    "Class 1": [
        ("cahier", "Book 80 leaves plain lines (Sc/Tech)", 3, ""),
        ("cahier", "Book 80 leaves square lines (Maths)", 2, ""),
        ("cahier", "Book 80 leaves plain lines (Eng/SWB)", 3, ""),
        ("cahier", "Book 60 leaves plain lines (S/S)", 2, ""),
        ("cahier", "Book 80 leaves plain lines (V/S)", 1, ""),
        ("cahier", "Book 20 leaves plain lines (Sports)", 1, ""),
        ("cahier", "Book 20 leaves red and blue lines (Writing)", 1, ""),
        ("cahier", "Book 20 leaves plain lines (Music)", 1, ""),
        ("cahier", "Cahier bananier", 6, ""),
        ("cahier", "Book 80 leaves plain lines (Home work)", 2, ""),
        ("cahier", "Book 40 leaves plain lines (ICT/NLC)", 2, ""),
        ("cahier", "Drawing book A4", 1, ""),
        ("cahier", "Textbook to indicate home work", 1, ""),
        ("ecriture", "Stylo bleu", 5, ""),
        ("ecriture", "Stylo noir", 3, ""),
        ("ecriture", "Stylo rouge", 2, ""),
        ("ecriture", "Sharpener", 5, ""),
        ("ecriture", "Eraser", 10, ""),
        ("ecriture", "UHU glue tube", 1, ""),
        ("ecriture", "Packet bold markers (BIC Velleda)", 2, ""),
        ("ecriture", "Packet colour pencils", 1, ""),
        ("ecriture", "Packet 2B or HB pencils", 2, ""),
        ("ecriture", "Packet patafis", 1, ""),
        ("ardoise", "White board", 1, ""),
        ("ardoise", "Duster", 1, ""),
        ("geometrie", "Ruler 30cm", 2, ""),
        ("rangement", "Plastic file", 1, ""),
        ("protection", "Transparent book covers", 20, "rouge, bleu, vert, jaune"),
        ("papier", "Packet canson paper A3 white", 1, ""),
    ],
    "Class 2": [
        ("cahier", "Book 80 leaves plain lines (English/SWB)", 3, ""),
        ("cahier", "Book 80 leaves square lines (Maths)", 2, ""),
        ("cahier", "Cahier de français 120 pages", 6, ""),
        ("cahier", "Book 80 leaves plain lines (S/T)", 2, ""),
        ("cahier", "Book 60 leaves plain lines (S/S)", 2, ""),
        ("cahier", "Book 80 leaves plain lines (V/S)", 1, ""),
        ("cahier", "Book 60 leaves plain lines (ICT/NLC)", 2, ""),
        ("cahier", "Book 20 leaves plain lines (Sports)", 1, ""),
        ("cahier", "Book 20 leaves plain lines (Music)", 1, ""),
        ("cahier", "Book 20 leaves red and blue lines (Writing)", 1, ""),
        ("cahier", "Textbook home work", 1, ""),
        ("cahier", "Drawing book A4", 1, ""),
        ("ecriture", "Stylo bleu", 5, ""),
        ("ecriture", "Stylo noir", 3, ""),
        ("ecriture", "Stylo rouge", 2, ""),
        ("ecriture", "Sharpener", 1, ""),
        ("ecriture", "Eraser", 5, ""),
        ("ecriture", "UHU glue stick and tube", 1, ""),
        ("ecriture", "Packet board markers", 1, ""),
        ("ecriture", "Packet colour pencils", 1, ""),
        ("ecriture", "Packet 2B or HB pencils", 1, ""),
        ("ecriture", "Packet patafis", 1, ""),
        ("ardoise", "White board", 1, ""),
        ("ardoise", "Duster", 1, ""),
        ("geometrie", "Ruler 30cm", 2, ""),
        ("rangement", "Plastic file", 1, ""),
        ("protection", "Transparent book cover", 20, "rouge, bleu, jaune, vert"),
        ("papier", "Packet canson paper A3", 1, ""),
    ],
    "Class 3": [
        ("cahier", "Book 80 leaves plain lines (English/Evaluation)", 3, ""),
        ("cahier", "Book 80 leaves square lines (Maths)", 2, ""),
        ("cahier", "Book 60 leaves plain lines (matières majeures)", 10, "S/W building, Health, Environmental, Tech, ICT, History, Geography, Home Ec, Peace"),
        ("cahier", "Book 40 leaves plain lines (matières mineures)", 8, "Civics, NLC, Human R, Agro Past, Arts, Moral Ed, Music"),
        ("cahier", "Book red and blue lines (writing)", 2, ""),
        ("cahier", "Cahier de français 144 pages", 2, ""),
        ("cahier", "Textbook home work", 1, ""),
        ("cahier", "Drawing book A4", 1, ""),
        ("ecriture", "Stylo bleu", 3, ""),
        ("ecriture", "Stylo noir", 1, ""),
        ("ecriture", "Stylo rouge", 2, ""),
        ("ecriture", "Sharpener", 1, ""),
        ("ecriture", "Eraser", 3, ""),
        ("ecriture", "UHU glue stick and tube", 1, ""),
        ("ecriture", "Packet bold markers", 1, ""),
        ("ecriture", "Packet colour pencils", 1, ""),
        ("ecriture", "Packet pencils 2B or HB", 1, ""),
        ("ardoise", "White board", 1, ""),
        ("ardoise", "Duster", 1, ""),
        ("geometrie", "Ruler 30cm", 1, ""),
        ("geometrie", "Mathematical set", 1, ""),
        ("rangement", "Plastic file", 1, ""),
        ("protection", "Transparent book cover", 20, "rouge, bleu, jaune, vert"),
    ],
    "Class 4": [
        ("cahier", "Book 80 leaves plain lines (English, History, Geography, Eval, SWB)", 6, ""),
        ("cahier", "Book 80 leaves square lines (Maths)", 2, ""),
        ("cahier", "Book 60 leaves plain lines (matières majeures)", 9, "Health, Environmental, ICT, Home Ec, Civics, Tech, Music, Agro-Past, Arts"),
        ("cahier", "Book 40 leaves plain lines (matières mineures)", 5, "Human Rights, Moral Ed, NLC, Sports, SWB"),
        ("cahier", "Cahier de français 144 pages", 2, ""),
        ("cahier", "Textbook home work", 1, ""),
        ("cahier", "Drawing book A4", 1, ""),
        ("ecriture", "Stylo bleu", 3, ""),
        ("ecriture", "Stylo rouge", 2, ""),
        ("ecriture", "Stylo noir", 1, ""),
        ("ecriture", "USB key 2GB", 1, ""),
        ("ecriture", "UHU glue stick and tube", 1, ""),
        ("ecriture", "Eraser", 3, ""),
        ("ecriture", "Sharpener", 1, ""),
        ("ecriture", "Packet board markers", 1, ""),
        ("ecriture", "Packet colour crayons", 1, ""),
        ("ecriture", "Pencils 2B or HB", 2, ""),
        ("ardoise", "White board", 1, ""),
        ("ardoise", "Duster", 1, ""),
        ("geometrie", "Ruler 30cm", 1, ""),
        ("geometrie", "Mathematical set", 1, ""),
        ("rangement", "Plastic file", 1, ""),
    ],
    "Class 5": [
        ("cahier", "Book 200 leaves ledger plain lines (English)", 2, ""),
        ("cahier", "Book 200 leaves ledger square lines (Maths)", 2, ""),
        ("cahier", "Book 80 leaves plain lines (matières majeures)", 10, "History, Geography, Health, Dictation, Eval, Tech, ICT, Home Ec"),
        ("cahier", "Book 60 leaves plain lines (matières mineures)", 6, "Civics, Agro Pastoral, Arts, SWB, Music, Sports"),
        ("cahier", "Book 40 leaves plain lines", 5, "Human Right, Moral Ed, visual Arts, NLC"),
        ("cahier", "Cahier de français 144 pages", 2, ""),
        ("cahier", "Drawing book A4", 1, ""),
        ("ecriture", "Stylo bleu", 3, ""),
        ("ecriture", "Stylo rouge", 2, ""),
        ("ecriture", "Stylo noir", 1, ""),
        ("ecriture", "USB key 2GB", 1, ""),
        ("ecriture", "UHU glue stick and tube", 1, ""),
        ("ecriture", "Sharpener", 1, ""),
        ("ecriture", "Packet board markers", 1, ""),
        ("ecriture", "Packet colour crayons", 1, ""),
        ("ecriture", "Pencils 2B or HB", 2, ""),
        ("ecriture", "Eraser", 1, ""),
        ("ardoise", "White board", 1, ""),
        ("ardoise", "Duster", 1, ""),
        ("geometrie", "Ruler 30cm", 1, ""),
        ("geometrie", "Mathematics set", 1, ""),
        ("rangement", "Pencil case", 1, ""),
        ("protection", "Transparent book cover", 20, "rouge, bleu, jaune, vert"),
    ],
    "Class 6": [
        ("cahier", "Book 200 leaves ledger plain lines (English)", 2, ""),
        ("cahier", "Book 200 leaves ledger square lines (Maths)", 2, ""),
        ("cahier", "Book 80 leaves plain lines (matières majeures)", 10, "Eval, Health, Environmental, SWB, History, Geography, ICT, Dictation, Tech"),
        ("cahier", "Book 60 leaves plain lines (matières mineures)", 6, "Home Ec, Agro Pastoral, Civics, Sports, Music, NLC"),
        ("cahier", "Book 40 leaves plain lines", 3, "Moral Ed, Human Rights, Arts & Craft"),
        ("cahier", "Cahier de français 144 pages", 2, ""),
        ("cahier", "Drawing book A4", 1, ""),
        ("ecriture", "Stylo bleu", 3, ""),
        ("ecriture", "Stylo rouge", 2, ""),
        ("ecriture", "Stylo noir", 1, ""),
        ("ecriture", "USB key", 1, ""),
        ("ecriture", "UHU glue stick and tube", 1, ""),
        ("ecriture", "Sharpener", 2, ""),
        ("ecriture", "Packet board markers", 1, ""),
        ("ecriture", "Packet colour pencils", 1, ""),
        ("ecriture", "Pencil 2B or HB", 1, ""),
        ("ecriture", "Eraser", 2, ""),
        ("ardoise", "White board", 1, ""),
        ("ardoise", "Duster", 1, ""),
        ("geometrie", "Ruler 30cm", 1, ""),
        ("geometrie", "Mathematics set", 1, ""),
        ("rangement", "Plastic file", 1, ""),
        ("protection", "Book covers", 20, "5 red, 5 blue, 5 yellow, 5 green"),
        ("protection", "Transparent book cover", 20, "red, blue, yellow"),
        ("dictionnaire", "Mac Milan school dictionary", 1, ""),
    ],
    # ────────── SECONDAIRE FRANCOPHONE 1er cycle ──────────
    "6ème": [
        ("cahier", "Cahier 288 pages (Langue française + Littérature)", 2, ""),
        ("cahier", "Cahier 120 pages (Langue française + Littérature)", 2, ""),
        ("cahier", "Cahier 200 pages ledgers (Anglais)", 1, ""),
        ("cahier", "Cahier 400 pages (Histoire + Géographie)", 1, ""),
        ("cahier", "Cahier 200 pages (Citoyenneté)", 1, ""),
        ("cahier", "Cahier 400 pages sans format (Mathématiques)", 1, ""),
        ("cahier", "Cahier 200 pages (Mathématiques)", 1, ""),
        ("cahier", "Cahier TP 300 pages (Sciences)", 1, ""),
        ("cahier", "Cahier 300 pages (Informatique)", 1, ""),
        ("cahier", "Cahier 144 pages petit format (Sport)", 1, ""),
        ("cahier", "Cahier 100 pages (TM)", 1, ""),
        ("ecriture", "Stylo bleu", 2, ""),
        ("ecriture", "Stylo rouge", 2, ""),
        ("ecriture", "Crayon ordinaire", 2, ""),
        ("ecriture", "Gomme blanche", 2, ""),
        ("ecriture", "Paquet crayons couleur", 1, ""),
        ("geometrie", "Compas", 1, ""),
        ("geometrie", "Rapporteur", 1, ""),
        ("geometrie", "Équerre", 1, ""),
        ("geometrie", "Règle 30cm", 1, ""),
        ("dictionnaire", "Dictionnaire Maxipoche Larousse", 1, ""),
        ("dictionnaire", "Dictionnaire français/anglais Larousse", 1, ""),
    ],
    "5ème": [
        ("cahier", "Cahier 288 pages (Langue française + Littérature)", 2, ""),
        ("cahier", "Cahier 120 pages (Langue française + Littérature)", 2, ""),
        ("cahier", "Cahier 200 pages ledgers (Anglais)", 1, ""),
        ("cahier", "Cahier 400 pages (Histoire + Géographie)", 1, ""),
        ("cahier", "Cahier 200 pages (Citoyenneté)", 1, ""),
        ("cahier", "Cahier 400 pages sans format (Mathématiques)", 1, ""),
        ("cahier", "Cahier 200 pages (Mathématiques)", 1, ""),
        ("cahier", "Cahier TP 300 pages (Sciences)", 1, ""),
        ("cahier", "Cahier 300 pages (Informatique)", 1, ""),
        ("cahier", "Cahier 144 pages (Sport)", 1, ""),
        ("cahier", "Cahier 100 pages (TM)", 1, ""),
        ("ecriture", "Stylo bleu", 2, ""),
        ("ecriture", "Stylo rouge", 2, ""),
        ("ecriture", "Crayon ordinaire", 2, ""),
        ("ecriture", "Gomme blanche", 2, ""),
        ("ecriture", "Paquet crayons couleur", 1, ""),
        ("geometrie", "Compas", 1, ""),
        ("geometrie", "Rapporteur", 1, ""),
        ("geometrie", "Équerre", 1, ""),
        ("geometrie", "Règle 30cm", 1, ""),
        ("dictionnaire", "Dictionnaire Maxipoche Larousse", 1, ""),
        ("dictionnaire", "Dictionnaire français/anglais Larousse", 1, ""),
    ],
    "4ème": [
        ("cahier", "Cahier 288 pages (Langue française + Littérature)", 2, ""),
        ("cahier", "Cahier 120 pages (Langue française + Littérature)", 2, ""),
        ("cahier", "Cahier 200 pages ledgers (Anglais)", 1, ""),
        ("cahier", "Cahier 400 pages (Histoire + Géographie)", 1, ""),
        ("cahier", "Cahier 200 pages (Citoyenneté)", 1, ""),
        ("cahier", "Cahier 300 pages (Allemand)", 1, ""),
        ("cahier", "Cahier 192 pages petit format (Espagnol)", 1, ""),
        ("cahier", "Cahier 400 pages sans format (Mathématiques)", 1, ""),
        ("cahier", "Cahier 200 pages (Mathématiques)", 1, ""),
        ("cahier", "Cahier TP 300 pages (SVTEEHB)", 1, ""),
        ("cahier", "Cahier 200 pages (SVTEEHB)", 1, ""),
        ("cahier", "Cahier 300 pages grand format (Physique/Chimie/Tech)", 1, ""),
        ("cahier", "Cahier TP 280 pages petit format (PCT)", 1, ""),
        ("cahier", "Cahier 300 pages (Informatique)", 1, ""),
        ("cahier", "Cahier 144 pages (Sport)", 1, ""),
        ("cahier", "Cahier 100 pages (TM)", 1, ""),
        ("ecriture", "Stylo bleu", 2, ""),
        ("ecriture", "Stylo rouge", 2, ""),
        ("ecriture", "Crayon ordinaire", 2, ""),
        ("ecriture", "Gomme blanche", 2, ""),
        ("ecriture", "Paquet crayons couleur", 1, ""),
        ("geometrie", "Compas", 1, ""),
        ("geometrie", "Rapporteur", 1, ""),
        ("geometrie", "Équerre 45°", 1, ""),
        ("geometrie", "Équerre 60°", 1, ""),
        ("geometrie", "Règle 30cm", 1, ""),
        ("dictionnaire", "Dictionnaire Maxipoche Larousse", 1, ""),
        ("dictionnaire", "Dictionnaire français/anglais Larousse", 1, ""),
    ],
    "3ème": [
        ("cahier", "Cahier 288 pages (Langue française + Littérature)", 2, ""),
        ("cahier", "Cahier 120 pages (Langue française + Littérature)", 2, ""),
        ("cahier", "Cahier 200 pages ledgers (Anglais)", 1, ""),
        ("cahier", "Cahier 400 pages (Histoire + Géographie)", 1, ""),
        ("cahier", "Cahier 200 pages (Citoyenneté)", 1, ""),
        ("cahier", "Cahier 300 pages (Allemand)", 1, ""),
        ("cahier", "Cahier 192 pages petit format (Espagnol)", 1, ""),
        ("cahier", "Cahier 400 pages sans format (Mathématiques)", 1, ""),
        ("cahier", "Cahier 200 pages (Mathématiques)", 1, ""),
        ("cahier", "Cahier TP 300 pages (SVTEEHB)", 1, ""),
        ("cahier", "Cahier 200 pages (SVTEEHB)", 1, ""),
        ("cahier", "Cahier 300 pages petit format (PCT)", 1, ""),
        ("cahier", "Cahier 300 pages grand format quadrillé (PCT)", 1, ""),
        ("cahier", "Cahier 300 pages (Informatique)", 1, ""),
        ("cahier", "Cahier 144 pages (Sport)", 1, ""),
        ("cahier", "Cahier 100 pages (TM)", 1, ""),
        ("ecriture", "Stylo bleu", 2, ""),
        ("ecriture", "Stylo rouge", 2, ""),
        ("ecriture", "Crayon ordinaire", 2, ""),
        ("ecriture", "Gomme blanche", 2, ""),
        ("ecriture", "Paquet crayons couleur", 1, ""),
        ("geometrie", "Compas", 1, ""),
        ("geometrie", "Rapporteur", 1, ""),
        ("geometrie", "Équerre 45°", 1, ""),
        ("geometrie", "Équerre 60°", 1, ""),
        ("geometrie", "Règle 30cm", 1, ""),
        ("geometrie", "Calculatrice scientifique", 1, ""),
        ("dictionnaire", "Dictionnaire Maxipoche Larousse", 1, ""),
        ("dictionnaire", "Dictionnaire français/anglais Larousse", 1, ""),
    ],
    # ────────── SECONDAIRE FRANCOPHONE 2nd cycle ──────────
    "2nde A-C": [
        ("cahier", "Cahier 288 pages (Langue française)", 1, ""),
        ("cahier", "Cahier 120 pages (Langue française)", 2, ""),
        ("cahier", "Cahier 300 pages ledgers (Anglais)", 1, ""),
        ("cahier", "Cahier 400 pages (Histoire)", 1, ""),
        ("cahier", "Cahier 200 pages (Citoyenneté)", 1, ""),
        ("cahier", "Cahier 400 pages sans format (Mathématiques)", 1, ""),
        ("cahier", "Cahier 100 pages (Mathématiques 2nde A4)", 1, ""),
        ("cahier", "Cahier TP 300 pages (Sciences 2nde A4)", 1, ""),
        ("cahier", "Cahier 300 pages (Informatique)", 1, ""),
        ("cahier", "Cahier 300 pages (Allemand/Espagnol)", 1, ""),
        ("cahier", "Cahier 400 pages sans format (Mathématiques 2nde C)", 1, ""),
        ("cahier", "Cahier 300 pages (Mathématiques 2nde C)", 1, ""),
        ("cahier", "Cahier TP 300 pages (SVTEEHB 2nde C)", 1, ""),
        ("cahier", "Cahier 200 pages (SVTEEHB 2nde C)", 1, ""),
        ("cahier", "Cahier 300 pages petit format (Physique-Chimie)", 1, ""),
        ("cahier", "Cahier 300 pages grand format quadrillé (Physique-Chimie)", 1, ""),
        ("cahier", "Cahier 144 pages (Sport)", 1, ""),
        ("cahier", "Cahier 100 pages (TM)", 1, ""),
        ("ecriture", "Stylo bleu", 2, ""),
        ("ecriture", "Stylo rouge", 2, ""),
        ("ecriture", "Crayon ordinaire", 2, ""),
        ("ecriture", "Gomme blanche", 2, ""),
        ("ecriture", "Boîte de colle", 1, ""),
        ("ecriture", "Paquet crayons couleur", 1, ""),
        ("geometrie", "Compas", 1, ""),
        ("geometrie", "Rapporteur", 1, ""),
        ("geometrie", "Équerre", 1, ""),
        ("geometrie", "Règle 30cm", 1, ""),
        ("geometrie", "Calculatrice scientifique", 1, ""),
        ("dictionnaire", "Dictionnaire Maxipoche Larousse", 1, ""),
        ("dictionnaire", "Dictionnaire français/anglais Larousse", 1, ""),
    ],
    "2nde STT": [
        ("cahier", "Cahier 288 pages (Français)", 1, ""),
        ("cahier", "Cahier 120 pages (Français)", 1, ""),
        ("cahier", "Cahier 300 pages ledgers (Anglais)", 1, ""),
        ("cahier", "Cahier 300 pages (Histoire)", 1, ""),
        ("cahier", "Cahier 400 pages (Géographie)", 1, ""),
        ("cahier", "Cahier 200 pages (Citoyenneté)", 1, ""),
        ("cahier", "Cahier 400 pages sans format (Maths générales)", 1, ""),
        ("cahier", "Cahier 300 pages (Maths générales)", 1, ""),
        ("cahier", "Cahier 300 pages (Informatique)", 1, ""),
        ("cahier", "Cahier 100 pages (Informatique)", 1, ""),
        ("cahier", "Cahier 200 pages (Comptabilité financière)", 1, ""),
        ("cahier", "Cahier 200 pages (Bureautique)", 1, ""),
        ("cahier", "Cahier 200 pages (Communication professionnelle)", 1, ""),
        ("cahier", "Cahier 200 pages (Commerce)", 1, ""),
        ("cahier", "Cahier 200 pages (Org. travail administratif)", 1, ""),
        ("cahier", "Cahier 200 pages (Économie générale)", 1, ""),
        ("cahier", "Cahier 200 pages (Maths/Stats appliquées)", 1, ""),
        ("cahier", "Cahier 144 pages (Sport)", 1, ""),
        ("cahier", "Cahier 100 pages (TM)", 1, ""),
        ("ecriture", "Stylo bleu", 2, ""),
        ("ecriture", "Stylo rouge", 2, ""),
        ("ecriture", "Crayon ordinaire", 2, ""),
        ("ecriture", "Gomme blanche", 2, ""),
        ("ecriture", "Boîte de colle", 1, ""),
        ("ecriture", "Paquet crayons couleur", 1, ""),
        ("geometrie", "Règle 30cm", 1, ""),
        ("geometrie", "Calculatrice financière", 1, ""),
        ("dictionnaire", "Dictionnaire Maxipoche Larousse", 1, ""),
        ("dictionnaire", "Dictionnaire français/anglais Larousse", 1, ""),
    ],
    "1ère": [
        ("cahier", "Cahier 288 pages (Langue française)", 1, ""),
        ("cahier", "Cahier 120 pages (Langue française)", 2, ""),
        ("cahier", "Cahier 300 pages ledgers (Anglais)", 1, ""),
        ("cahier", "Cahier 400 pages (Histoire)", 1, ""),
        ("cahier", "Cahier 200 pages (Citoyenneté)", 1, ""),
        ("cahier", "Cahier 400 pages grand format (Mathématiques 1ère C)", 1, ""),
        ("cahier", "Cahier 100 pages (Mathématiques 1ère C)", 1, ""),
        ("cahier", "Cahier 400 pages grand format (Mathématiques 1ère D)", 1, ""),
        ("cahier", "Cahier 300 pages (Mathématiques 1ère D)", 1, ""),
        ("cahier", "Cahier TP 300 pages (SVTEEHB)", 1, ""),
        ("cahier", "Cahier 200 pages (SVTEEHB)", 1, ""),
        ("cahier", "Cahier 300 pages grand format (Physique)", 1, ""),
        ("cahier", "Cahier 200 pages petit format (Physique)", 1, ""),
        ("cahier", "Cahier 300 pages petit format (Chimie)", 1, ""),
        ("cahier", "Cahier 200 pages petit format (Chimie)", 1, ""),
        ("cahier", "Cahier 300 pages (Informatique)", 1, ""),
        ("cahier", "Cahier 100 pages (Informatique 1ère D)", 1, ""),
        ("cahier", "Cahier 144 pages (Sport)", 1, ""),
        ("cahier", "Cahier 100 pages (TM)", 1, ""),
        ("ecriture", "Stylo bleu", 2, ""),
        ("ecriture", "Stylo rouge", 2, ""),
        ("ecriture", "Crayon ordinaire", 2, ""),
        ("ecriture", "Gomme blanche", 2, ""),
        ("ecriture", "Boîte de colle", 1, ""),
        ("ecriture", "Paquet crayons couleur", 1, ""),
        ("geometrie", "Compas", 1, ""),
        ("geometrie", "Rapporteur", 1, ""),
        ("geometrie", "Équerre 45°", 1, ""),
        ("geometrie", "Équerre 60°", 1, ""),
        ("geometrie", "Règle 30cm", 1, ""),
        ("geometrie", "Calculatrice scientifique", 1, ""),
        ("dictionnaire", "Dictionnaire Maxipoche Larousse", 1, ""),
        ("dictionnaire", "Dictionnaire français/anglais Larousse", 1, ""),
    ],
    # ────────── SECONDAIRE ANGLOPHONE ──────────
    "Form 1": [
        ("cahier", "Plain lines 200 leaves Ledger A4 hard cover (English Language)", 1, ""),
        ("cahier", "Plain lines 300 leaves Ledger A4 hard cover (French)", 1, ""),
        ("cahier", "Plain lines 200 leaves Ledger A4 hard cover (Biology)", 1, ""),
        ("cahier", "Plain lines 200 leaves Ledger A4 hard cover (Chemistry)", 1, ""),
        ("cahier", "Plain lines 200 leaves Ledger A4 hard cover (Geography)", 1, ""),
        ("cahier", "Plain lines 80 leaves A5 (Geography)", 1, ""),
        ("cahier", "Graph book (Geography)", 1, ""),
        ("cahier", "Plain lines 200 leaves Ledger A4 hard cover (History)", 1, ""),
        ("cahier", "Plain lines 80 leaves A5 (History)", 1, ""),
        ("cahier", "Plain lines 200 leaves Ledger A4 hard cover (Physics)", 1, ""),
        ("cahier", "Squared 200 leaves Ledger A4 hard cover (Mathematics)", 2, ""),
        ("cahier", "Plain lines 200 leaves Ledger A4 hard cover (Computer Science)", 1, ""),
        ("cahier", "Plain line 80 leaves (Food and Nutrition)", 1, ""),
        ("cahier", "Plain lines 200 leaves Ledger A4 hard cover (Literature)", 1, ""),
        ("cahier", "Plain line 200 leaves Ledger A4 hard cover (Citizenship)", 1, ""),
        ("cahier", "Plain line 80 leaves A5 (Citizenship)", 1, ""),
        ("cahier", "Plain line 80 leaves (Sport)", 1, ""),
        ("cahier", "Plain line 80 leaves (Manual Labour)", 1, ""),
        ("cahier", "Plain lines 200 leaves Ledger A4 hard cover (Accounting)", 1, ""),
        ("cahier", "Plain lines 200 leaves Ledger A4 hard cover (Marketing)", 1, ""),
        ("cahier", "Plain lines 200 leaves Ledger A4 hard cover (Office Practice)", 1, ""),
        ("ecriture", "Stylo bleu", 3, ""),
        ("ecriture", "Stylo rouge", 2, ""),
        ("ecriture", "Stylo noir", 1, ""),
        ("ecriture", "Rotring pencil", 1, ""),
        ("ecriture", "Crayon ordinaire", 2, ""),
        ("ecriture", "Gomme", 2, ""),
        ("ecriture", "Packet colour pencils", 1, ""),
        ("geometrie", "Ruler 30cm", 1, ""),
        ("geometrie", "Mathematical set", 1, ""),
        ("geometrie", "Scientific calculator", 1, ""),
        ("dictionnaire", "English dictionary", 1, ""),
        ("dictionnaire", "French dictionary", 1, ""),
    ],
    "Form 2": [
        ("cahier", "Plain lines 200 leaves Ledger A4 hard cover (English Language)", 1, ""),
        ("cahier", "Plain lines 300 leaves Ledger A4 hard cover (French)", 1, ""),
        ("cahier", "Plain lines 200 leaves Ledger A4 hard cover (Biology)", 1, ""),
        ("cahier", "Plain lines 200 leaves Ledger A4 hard cover (Chemistry)", 1, ""),
        ("cahier", "Plain lines 200 leaves Ledger A4 (Geography)", 1, ""),
        ("cahier", "Plain lines 80 leaves A5 (Geography)", 1, ""),
        ("cahier", "Graph book (Geography)", 1, ""),
        ("cahier", "Plain lines 200 leaves Ledger A4 (History)", 1, ""),
        ("cahier", "Plain lines 80 leaves A5 (History)", 1, ""),
        ("cahier", "Plain lines 200 leaves Ledger A4 (Physics)", 1, ""),
        ("cahier", "Squared 200 leaves Ledger A4 (Mathematics)", 2, ""),
        ("cahier", "Plain lines 200 leaves Ledger A4 (Computer Science)", 1, ""),
        ("cahier", "Plain line 80 leaves (Food and Nutrition)", 1, ""),
        ("cahier", "Plain lines 200 leaves Ledger A4 (Literature)", 1, ""),
        ("cahier", "Plain line 200 leaves Ledger A4 (Citizenship)", 1, ""),
        ("cahier", "Plain line 80 leaves A5 (Citizenship)", 1, ""),
        ("cahier", "Plain line 80 leaves (Sport)", 1, ""),
        ("cahier", "Plain line 80 leaves (Manual Labour)", 1, ""),
        ("cahier", "Plain lines 200 leaves Ledger A4 (Accounting)", 1, ""),
        ("cahier", "Plain lines 200 leaves Ledger A4 (Marketing)", 1, ""),
        ("cahier", "Plain lines 200 leaves Ledger A4 (Office Practice)", 1, ""),
        ("ecriture", "Stylo bleu", 3, ""),
        ("ecriture", "Stylo rouge", 2, ""),
        ("ecriture", "Stylo noir", 1, ""),
        ("ecriture", "Crayon ordinaire", 2, ""),
        ("ecriture", "Gomme", 2, ""),
        ("ecriture", "Packet colour pencils", 1, ""),
        ("geometrie", "Ruler 30cm", 1, ""),
        ("geometrie", "Mathematical set", 1, ""),
        ("geometrie", "Scientific calculator", 1, ""),
        ("dictionnaire", "English dictionary", 1, ""),
        ("dictionnaire", "French dictionary", 1, ""),
    ],
    "Form 3 Commercial": [
        ("cahier", "Plain lines 200 leaves Ledger A4 (English)", 1, ""),
        ("cahier", "Plain lines ledger 300 pages (French)", 1, ""),
        ("cahier", "Plain lines 500 leaves Ledger A4 (Geography)", 1, ""),
        ("cahier", "Plain lines 80 leaves A5 (Geography)", 1, ""),
        ("cahier", "Ledger 300 pages square lines (Maths)", 1, ""),
        ("cahier", "Graph book (Maths)", 1, ""),
        ("cahier", "Plain lines 80 leaves A5 (Maths)", 1, ""),
        ("cahier", "300 pages ledger (Computer Science)", 1, ""),
        ("cahier", "Plain lines ledger 300 pages (Economics)", 1, ""),
        ("cahier", "Plain line 300 leaves Ledger A4 (Citizenship)", 1, ""),
        ("cahier", "Plain line 80 leaves A5 (Citizenship)", 1, ""),
        ("cahier", "Plain lines 80 leaves (Sport)", 1, ""),
        ("cahier", "Plain lines 80 leaves (Manual Labour)", 1, ""),
        ("cahier", "Plain lines 200 leaves Ledger A4 (Accounting)", 1, ""),
        ("cahier", "Plain lines 200 leaves Ledger A4 (Marketing)", 1, ""),
        ("cahier", "Plain lines 200 leaves Ledger A4 (Office Practice)", 1, ""),
        ("cahier", "Plain lines 200 leaves Ledger A4 (Business Maths)", 2, ""),
        ("ecriture", "Stylo bleu", 3, ""),
        ("ecriture", "Stylo rouge", 2, ""),
        ("ecriture", "Stylo noir", 1, ""),
        ("ecriture", "Crayon ordinaire", 2, ""),
        ("ecriture", "Gomme", 2, ""),
        ("geometrie", "Ruler 30cm", 1, ""),
        ("geometrie", "Financial calculator", 1, ""),
        ("dictionnaire", "English dictionary", 1, ""),
        ("dictionnaire", "French dictionary", 1, ""),
    ],
    "Form 4 Commercial": [
        ("cahier", "Plain lines 200 leaves Ledger A4 (English)", 1, ""),
        ("cahier", "Plain lines ledger 300 pages (French)", 1, ""),
        ("cahier", "Plain lines 500 leaves Ledger A4 (Geography)", 1, ""),
        ("cahier", "Plain lines 80 leaves A5 (Geography)", 1, ""),
        ("cahier", "Ledger 300 pages square lines (Maths)", 1, ""),
        ("cahier", "Graph book (Maths)", 1, ""),
        ("cahier", "300 pages ledger (Computer Science)", 1, ""),
        ("cahier", "Plain lines ledger 300 pages (Economics)", 1, ""),
        ("cahier", "Plain line 300 leaves Ledger A4 (Citizenship)", 1, ""),
        ("cahier", "Plain lines 200 leaves Ledger A4 (Accounting)", 1, ""),
        ("cahier", "Plain lines 200 leaves Ledger A4 (Marketing)", 1, ""),
        ("cahier", "Plain lines 200 leaves Ledger A4 (Office Practice)", 1, ""),
        ("cahier", "Plain lines 200 leaves Ledger A4 (Business Maths)", 2, ""),
        ("ecriture", "Stylo bleu", 3, ""),
        ("ecriture", "Stylo rouge", 2, ""),
        ("ecriture", "Stylo noir", 1, ""),
        ("ecriture", "Crayon ordinaire", 2, ""),
        ("ecriture", "Gomme", 2, ""),
        ("geometrie", "Ruler 30cm", 1, ""),
        ("geometrie", "Financial calculator", 1, ""),
        ("dictionnaire", "English dictionary", 1, ""),
        ("dictionnaire", "French dictionary", 1, ""),
    ],
    "Form 4 General": [
        ("cahier", "Plain lines 200 leaves Ledger A4 (English Language)", 1, ""),
        ("cahier", "Plain lines 200 leaves Ledger A4 (French)", 1, ""),
        ("cahier", "Plain lines 80 leaves (French)", 1, ""),
        ("cahier", "Plain lines Ledger 300 pages (Biology)", 1, ""),
        ("cahier", "Plain lines Ledger 200 pages (Human Biology)", 1, ""),
        ("cahier", "Plain lines Ledger 400 pages (Chemistry)", 1, ""),
        ("cahier", "Plain lines 500 leaves Ledger A4 (Geography)", 1, ""),
        ("cahier", "Plain lines 80 leaves A5 (Geography)", 1, ""),
        ("cahier", "Plain lines 500 leaves Ledger A4 (History)", 1, ""),
        ("cahier", "Plain lines 80 leaves A5 (History)", 1, ""),
        ("cahier", "Plain lines Ledger 200 pages (Physics)", 1, ""),
        ("cahier", "Square lines Ledger 300 pages (Maths)", 1, ""),
        ("cahier", "Square lines 80 leaves A5 (Maths)", 1, ""),
        ("cahier", "300 pages ledger (Computer Science)", 1, ""),
        ("cahier", "Plain lines Ledger 200 pages (Food and Nutrition)", 1, ""),
        ("cahier", "Plain lines 200 leaves Ledger A4 (Literature)", 1, ""),
        ("cahier", "Plain lines 300 leaves Ledger A4 (Citizenship)", 1, ""),
        ("cahier", "Plain lines 80 leaves A5 (Citizenship)", 1, ""),
        ("cahier", "Plain lines 80 leaves A5 (Sports)", 1, ""),
        ("cahier", "Plain lines Ledger 200 leaves (Accounting)", 1, ""),
        ("ecriture", "Stylo bleu", 3, ""),
        ("ecriture", "Stylo rouge", 2, ""),
        ("ecriture", "Stylo noir", 1, ""),
        ("ecriture", "Rotring pencil", 1, ""),
        ("ecriture", "Crayon ordinaire", 2, ""),
        ("ecriture", "Gomme", 2, ""),
        ("geometrie", "Ruler 30cm", 1, ""),
        ("geometrie", "Mathematical set", 1, ""),
        ("geometrie", "Scientific calculator", 1, ""),
        ("dictionnaire", "English dictionary", 1, ""),
        ("dictionnaire", "French dictionary", 1, ""),
    ],
    "Form 5": [
        ("cahier", "Plain lines 300 leaves Ledger A4 (English Language)", 1, ""),
        ("cahier", "Plain lines 300 leaves Ledger A4 (French)", 1, ""),
        ("cahier", "Plain lines 80 leaves (French)", 1, ""),
        ("cahier", "Plain lines Ledger 300 pages (Biology)", 1, ""),
        ("cahier", "Plain lines Ledger 300 pages (Human Biology)", 1, ""),
        ("cahier", "Plain lines 80 leaves (Human Biology)", 1, ""),
        ("cahier", "Plain lines Ledger 300 pages (Chemistry)", 1, ""),
        ("cahier", "Plain lines 80 leaves (Chemistry)", 1, ""),
        ("cahier", "Plain lines 500 leaves Ledger A4 (Geography)", 1, ""),
        ("cahier", "Plain lines 80 leaves A5 (Geography)", 1, ""),
        ("cahier", "Graph book (Geography)", 1, ""),
        ("cahier", "Rotring pencil (Geography)", 1, ""),
        ("cahier", "Plain lines 500 leaves Ledger A4 (History)", 1, ""),
        ("cahier", "Plain lines 80 leaves A5 (History)", 1, ""),
        ("cahier", "Plain lines Ledger 300 pages (Physics)", 1, ""),
        ("cahier", "Squared lines Ledger 300 pages (Maths)", 1, ""),
        ("cahier", "Plain lines 80 leaves (Maths)", 1, ""),
        ("cahier", "300 pages ledger (Computer Science)", 1, ""),
        ("cahier", "200 pages ledger (Food and Nutrition)", 1, ""),
        ("cahier", "Plain lines 300 leaves Ledger A4 (Literature)", 1, ""),
        ("cahier", "Plain lines 300 leaves Ledger A4 (Citizenship)", 1, ""),
        ("cahier", "Plain lines 80 leaves A5 (Citizenship)", 1, ""),
        ("cahier", "Plain lines 80 leaves (Sports)", 1, ""),
        ("cahier", "Plain lines Ledger 200 leaves (Accounting)", 2, ""),
        ("cahier", "Plain lines Ledger 200 pages (Logic)", 1, ""),
        ("ecriture", "Stylo bleu", 3, ""),
        ("ecriture", "Stylo rouge", 2, ""),
        ("ecriture", "Stylo noir", 1, ""),
        ("ecriture", "Crayon ordinaire", 2, ""),
        ("ecriture", "Gomme", 2, ""),
        ("geometrie", "Ruler 30cm", 1, ""),
        ("geometrie", "Mathematical set", 1, ""),
        ("geometrie", "Scientific calculator", 1, ""),
        ("dictionnaire", "English dictionary", 1, ""),
        ("dictionnaire", "French dictionary", 1, ""),
    ],
    # Lower Sixth et Upper Sixth partagent le même PDF (A-Level), donc même
    # liste de fournitures dupliquée pour matcher les 2 classes distinctes
    # de programmes_scolaires.
    "Lower Sixth": [
        # A-Level : pas de tableau exercise books détaillé, kit générique avancé
        ("cahier", "Plain lines 500 leaves Ledger A4 (English Language)", 1, ""),
        ("cahier", "Plain lines 500 leaves Ledger A4 (French)", 1, ""),
        ("cahier", "Plain lines 500 leaves Ledger A4 (Literature)", 1, ""),
        ("cahier", "Plain lines 500 leaves Ledger A4 (Geography)", 1, ""),
        ("cahier", "Plain lines 80 leaves A5 (Geography)", 1, ""),
        ("cahier", "Graph book (Geography)", 1, ""),
        ("cahier", "Plain lines 500 leaves Ledger A4 (History)", 1, ""),
        ("cahier", "Plain lines 500 leaves Ledger A4 (Philosophy)", 1, ""),
        ("cahier", "Plain lines Ledger 500 pages (Pure Mathematics)", 1, ""),
        ("cahier", "Plain lines Ledger 500 pages (Mechanics)", 1, ""),
        ("cahier", "Plain lines Ledger 500 pages (Chemistry)", 1, ""),
        ("cahier", "Plain lines Ledger 500 pages (Biology)", 1, ""),
        ("cahier", "Plain lines Ledger 500 pages (Physics)", 1, ""),
        ("cahier", "Plain lines Ledger 500 pages (ICT)", 1, ""),
        ("cahier", "Plain lines Ledger 500 pages (Computer Science)", 1, ""),
        ("cahier", "Plain lines Ledger 500 pages (Economics)", 1, ""),
        ("cahier", "Plain lines Ledger 500 pages (Statistics)", 1, ""),
        ("ecriture", "Stylo bleu", 3, ""),
        ("ecriture", "Stylo rouge", 2, ""),
        ("ecriture", "Stylo noir", 1, ""),
        ("ecriture", "Rotring pencil", 1, ""),
        ("ecriture", "Crayon ordinaire", 2, ""),
        ("ecriture", "Gomme", 2, ""),
        ("geometrie", "Ruler 30cm", 1, ""),
        ("geometrie", "Mathematical set", 1, ""),
        ("geometrie", "Scientific calculator", 1, ""),
        ("dictionnaire", "English dictionary", 1, ""),
        ("dictionnaire", "French dictionary", 1, ""),
    ],
}

# Upper Sixth = même contenu que Lower Sixth (PDF unique partagé A-Level)
DATA["Upper Sixth"] = DATA["Lower Sixth"]

# Mapping classe → (pays, systeme_educatif, niveau)
# Classes nommées comme dans programmes_scolaires (sans suffixe " Ord" car le
# système éducatif suffit à distinguer francophone ordinaire vs bilingue).
META = {
    # Primaire francophone ordinaire
    "SIL": ("Cameroun", "francophone", "Primaire"),
    "CP": ("Cameroun", "francophone", "Primaire"),
    "CE1": ("Cameroun", "francophone", "Primaire"),
    "CE2": ("Cameroun", "francophone", "Primaire"),
    "CM1": ("Cameroun", "francophone", "Primaire"),
    "CM2": ("Cameroun", "francophone", "Primaire"),
    # Primaire bilingue (NOUVELLES)
    "SIL BIL": ("Cameroun", "bilingue", "Primaire"),
    "CP BIL": ("Cameroun", "bilingue", "Primaire"),
    "CE1 BIL": ("Cameroun", "bilingue", "Primaire"),
    "CE2 BIL": ("Cameroun", "bilingue", "Primaire"),
    "CM1 BIL": ("Cameroun", "bilingue", "Primaire"),
    # Primaire anglophone
    "Class 1": ("Cameroun", "anglophone", "Primaire"),
    "Class 2": ("Cameroun", "anglophone", "Primaire"),
    "Class 3": ("Cameroun", "anglophone", "Primaire"),
    "Class 4": ("Cameroun", "anglophone", "Primaire"),
    "Class 5": ("Cameroun", "anglophone", "Primaire"),
    "Class 6": ("Cameroun", "anglophone", "Primaire"),
    # Secondaire francophone (niveau='Secondaire' : collège ET lycée offrent
    # généralement les deux cycles au Cameroun, cf. project_etab_secondaire_cm)
    "6ème": ("Cameroun", "francophone", "Secondaire"),
    "5ème": ("Cameroun", "francophone", "Secondaire"),
    "4ème": ("Cameroun", "francophone", "Secondaire"),
    "3ème": ("Cameroun", "francophone", "Secondaire"),
    "2nde A-C": ("Cameroun", "francophone", "Secondaire"),
    "2nde STT": ("Cameroun", "francophone", "Secondaire"),
    "1ère": ("Cameroun", "francophone", "Secondaire"),
    # Secondaire anglophone (idem : niveau='Secondaire')
    "Form 1": ("Cameroun", "anglophone", "Secondaire"),
    "Form 2": ("Cameroun", "anglophone", "Secondaire"),
    "Form 3 Commercial": ("Cameroun", "anglophone", "Secondaire"),
    "Form 4 Commercial": ("Cameroun", "anglophone", "Secondaire"),
    "Form 4 General": ("Cameroun", "anglophone", "Secondaire"),
    "Form 5": ("Cameroun", "anglophone", "Secondaire"),
    "Lower Sixth": ("Cameroun", "anglophone", "Secondaire"),
    "Upper Sixth": ("Cameroun", "anglophone", "Secondaire"),
}

CATEGORIE_LABELS = {
    "cahier": "Cahier",
    "ecriture": "Écriture",
    "geometrie": "Géométrie",
    "ardoise": "Ardoise",
    "protection": "Protection",
    "dictionnaire": "Dictionnaire",
    "papier": "Papier",
    "art": "Art",
    "rangement": "Rangement",
}


import re  # utilisé plus bas pour split_article


def split_article_canonique(article: str) -> str:
    """Renvoie l'article sans la partie entre parenthèses (= canonique)."""
    m = re.match(r"^(.+?)\s*\(([^)]+)\)\s*$", article)
    return m.group(1).strip() if m else article.strip()


def main():
    out_path = Path(__file__).parent / "referentiel_fournitures_par_classe.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Fournitures par classe"

    # En-têtes (avec clé masquée pour VLOOKUP depuis 'Articles uniques')
    headers = [
        "Pays", "Système", "Niveau", "Classe",
        "Catégorie", "Article", "Quantité", "Spécification",
        "Article canonique", "Clé",
        "Prix Standard (XAF)", "Prix Haut de Gamme (XAF)",
        "Sous-total Standard", "Sous-total Haut de Gamme",
    ]
    ws.append(headers)

    # Style en-têtes
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(border_style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border

    # Données (avec colonnes calculées E=Catégorie I=Canonique J=Clé +
    # formules VLOOKUP en K/L qui pointent vers 'Articles uniques')
    row_idx = 2
    for classe, items in DATA.items():
        pays, systeme, niveau = META[classe]
        for cat, article, qte, spec in items:
            canonique = split_article_canonique(article)
            ws.cell(row=row_idx, column=1, value=pays).border = border
            ws.cell(row=row_idx, column=2, value=systeme).border = border
            ws.cell(row=row_idx, column=3, value=niveau).border = border
            ws.cell(row=row_idx, column=4, value=classe).border = border
            ws.cell(row=row_idx, column=5, value=CATEGORIE_LABELS[cat]).border = border
            ws.cell(row=row_idx, column=6, value=article).border = border
            qte_cell = ws.cell(row=row_idx, column=7, value=qte)
            qte_cell.alignment = Alignment(horizontal="center")
            qte_cell.border = border
            ws.cell(row=row_idx, column=8, value=spec).border = border
            # Colonne 9 = Article canonique (utilisé pour la clé de lookup)
            ws.cell(row=row_idx, column=9, value=canonique).border = border
            # Colonne 10 = Clé concaténée : "Catégorie|Canonique|Spécification"
            ws.cell(
                row=row_idx, column=10,
                value=f'=E{row_idx}&"|"&I{row_idx}&"|"&H{row_idx}',
            ).border = border
            # Colonne 11 = Prix Standard (VLOOKUP depuis 'Articles uniques' col A → col G)
            ws.cell(
                row=row_idx, column=11,
                value=f"=IFERROR(VLOOKUP(J{row_idx},'Articles uniques'!A:H,7,FALSE),\"\")",
            ).border = border
            # Colonne 12 = Prix Haut de Gamme (VLOOKUP col A → col H)
            ws.cell(
                row=row_idx, column=12,
                value=f"=IFERROR(VLOOKUP(J{row_idx},'Articles uniques'!A:H,8,FALSE),\"\")",
            ).border = border
            # Colonne 13 = Sous-total Standard (Qté × Prix Std)
            ws.cell(
                row=row_idx, column=13,
                value=f'=IF(ISNUMBER(K{row_idx}),G{row_idx}*K{row_idx},"")',
            ).border = border
            # Colonne 14 = Sous-total Haut de Gamme (Qté × Prix HdG)
            ws.cell(
                row=row_idx, column=14,
                value=f'=IF(ISNUMBER(L{row_idx}),G{row_idx}*L{row_idx},"")',
            ).border = border
            row_idx += 1

    # Largeurs colonnes (colonnes 9 et 10 masquées : article canonique + clé interne)
    widths = {
        1: 10, 2: 12, 3: 10, 4: 18, 5: 14, 6: 50, 7: 9, 8: 24,
        9: 1, 10: 1,  # hidden
        11: 18, 12: 22, 13: 18, 14: 22,
    }
    for col_idx, w in widths.items():
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = w
    # Masque les colonnes I et J (canonique + clé technique)
    ws.column_dimensions["I"].hidden = True
    ws.column_dimensions["J"].hidden = True

    # Format devise FCFA (colonnes prix + sous-totaux)
    money_fmt = '#,##0" XAF"'
    for r in range(2, row_idx):
        for c in (11, 12, 13, 14):
            ws.cell(row=r, column=c).number_format = money_fmt

    # Fige en-tête + filtre
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:N{row_idx - 1}"

    # Feuille 2 : résumé par classe (qté totale par catégorie)
    ws2 = wb.create_sheet("Résumé par classe")
    summary_headers = ["Classe", "Pays", "Système", "Niveau", "Nb articles distincts", "Qté totale"]
    ws2.append(summary_headers)
    for col_idx in range(1, len(summary_headers) + 1):
        cell = ws2.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border

    row_idx = 2
    for classe, items in DATA.items():
        pays, systeme, niveau = META[classe]
        nb_articles = len(items)
        qte_totale = sum(qte for _, _, qte, _ in items)
        ws2.cell(row=row_idx, column=1, value=classe).border = border
        ws2.cell(row=row_idx, column=2, value=pays).border = border
        ws2.cell(row=row_idx, column=3, value=systeme).border = border
        ws2.cell(row=row_idx, column=4, value=niveau).border = border
        ws2.cell(row=row_idx, column=5, value=nb_articles).border = border
        ws2.cell(row=row_idx, column=6, value=qte_totale).border = border
        row_idx += 1

    summary_widths = {1: 22, 2: 12, 3: 14, 4: 12, 5: 20, 6: 14}
    for col_idx, w in summary_widths.items():
        ws2.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = w
    ws2.freeze_panes = "A2"

    # ─────────────────────────────────────────────────────────────────────────
    # Feuille 3 : Articles uniques (dédoublonnés) — pour saisir prix unitaire
    # ─────────────────────────────────────────────────────────────────────────
    # On dédoublonne par (catégorie, article_sans_matiere, spec). La "matière"
    # est ce qui se trouve entre parenthèses dans l'article.
    def split_article(article: str):
        """Sépare l'article en (canonique, matière entre parenthèses)."""
        m = re.match(r"^(.+?)\s*\(([^)]+)\)\s*$", article)
        if m:
            return m.group(1).strip(), m.group(2).strip()
        return article.strip(), ""

    aggregator = {}  # key = (categorie, canonique, specification) → {variantes, qte_totale, classes}
    for classe, items in DATA.items():
        for cat, article, qte, spec in items:
            canonique, matiere = split_article(article)
            key = (cat, canonique, spec)
            if key not in aggregator:
                aggregator[key] = {
                    "variantes": set(),
                    "qte_totale": 0,
                    "classes": set(),
                }
            if matiere:
                aggregator[key]["variantes"].add(matiere)
            aggregator[key]["qte_totale"] += qte
            aggregator[key]["classes"].add(classe)

    ws3 = wb.create_sheet("Articles uniques")
    # Colonne A = Clé concaténée (utilisée pour VLOOKUP depuis feuille 1)
    # Colonnes G/H = Prix Standard / Prix Haut de Gamme à remplir
    uniq_headers = [
        "Clé (ne pas modifier)",
        "Catégorie", "Article (canonique)", "Spécification / Couleur",
        "Matières / Variantes (info)", "Qté totale (toutes classes)",
        "Prix Standard (XAF)", "Prix Haut de Gamme (XAF)",
    ]
    ws3.append(uniq_headers)
    for col_idx in range(1, len(uniq_headers) + 1):
        cell = ws3.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border

    sorted_keys = sorted(aggregator.keys(), key=lambda k: (k[0], k[1], k[2]))
    row_idx = 2
    money_fmt = '#,##0" XAF"'
    for (cat, canonique, spec) in sorted_keys:
        data = aggregator[(cat, canonique, spec)]
        variantes = ", ".join(sorted(data["variantes"])) if data["variantes"] else ""
        cle = f"{CATEGORIE_LABELS[cat]}|{canonique}|{spec}"
        ws3.cell(row=row_idx, column=1, value=cle).border = border
        ws3.cell(row=row_idx, column=2, value=CATEGORIE_LABELS[cat]).border = border
        ws3.cell(row=row_idx, column=3, value=canonique).border = border
        ws3.cell(row=row_idx, column=4, value=spec).border = border
        ws3.cell(row=row_idx, column=5, value=variantes).border = border
        qte_cell = ws3.cell(row=row_idx, column=6, value=data["qte_totale"])
        qte_cell.alignment = Alignment(horizontal="center")
        qte_cell.border = border
        # Colonnes 7 et 8 = Prix Standard et Prix Haut de Gamme (à remplir)
        prix_std = ws3.cell(row=row_idx, column=7, value="")
        prix_std.border = border
        prix_std.number_format = money_fmt
        # Fond léger jaune pour signaler que ces cellules sont à remplir
        prix_std.fill = PatternFill(start_color="FFFBEB", end_color="FFFBEB", fill_type="solid")
        prix_hdg = ws3.cell(row=row_idx, column=8, value="")
        prix_hdg.border = border
        prix_hdg.number_format = money_fmt
        prix_hdg.fill = PatternFill(start_color="FFFBEB", end_color="FFFBEB", fill_type="solid")
        row_idx += 1

    # Colonne A (clé) masquée
    uniq_widths = {1: 1, 2: 14, 3: 55, 4: 32, 5: 50, 6: 16, 7: 20, 8: 24}
    for col_idx, w in uniq_widths.items():
        ws3.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = w
    ws3.column_dimensions["A"].hidden = True
    ws3.freeze_panes = "B2"
    ws3.auto_filter.ref = f"B1:H{row_idx - 1}"

    try:
        wb.save(out_path)
        excel_ok = True
    except PermissionError:
        excel_ok = False
        print(f"WARN : Excel ouvert/verrouille ({out_path}), regeneration ignoree")
    total_lines = sum(len(items) for items in DATA.values())
    nb_uniq = len(aggregator)
    if excel_ok:
        print(f"OK -> {out_path}")
        print(f"   {len(DATA)} classes, {total_lines} articles ventilés par classe")
        print(f"   {nb_uniq} articles UNIQUES (feuille 'Articles uniques')")

    # ─────────────────────────────────────────────────────────────────────────
    # Génère aussi un fichier SQL d'INSERT pour seeder
    # accessoires_populaires_par_classe (prod). Prix laissés à NULL — à
    # remplir plus tard quand le user aura rempli l'Excel.
    # ─────────────────────────────────────────────────────────────────────────
    sql_path = Path(__file__).parent / "seed_accessoires_populaires.sql"
    pays = "CM"
    annee_systeme = {
        "francophone": "CM-fr",
        "anglophone": "CM-en",
        "bilingue": "CM-bi",
    }

    def normalize(s: str) -> str:
        """Normalisation simple pour la déduplication (lower + trim + collapse)."""
        return re.sub(r"\s+", " ", s.strip().lower())

    lines_sql = []
    lines_sql.append(
        "-- =============================================================================\n"
        "-- SEED : accessoires_populaires_par_classe — fournitures (cahiers + accessoires)\n"
        "-- =============================================================================\n"
        "-- Date    : 2026-05-23\n"
        "-- Source  : 31 PDFs Soft Education 2025-2026 (extraction Claude)\n"
        "--\n"
        "-- 31 classes × N articles = 836 lignes. Prix laissés à NULL — à remplir\n"
        "-- après que l'utilisateur ait complété la feuille 'Articles uniques'\n"
        "-- de scripts/seed-fournitures/referentiel_fournitures_par_classe.xlsx.\n"
        "--\n"
        "-- Idempotent : ON CONFLICT sur l'unique (pays, niveau, classe, nom_normalise)\n"
        "-- met à jour quantite_mediane et increment occurrences si déjà présent.\n"
        "--\n"
        "-- Pas de BEGIN/COMMIT : sqlx::migrate!() gère la transaction.\n"
        "-- =============================================================================\n\n"
    )

    def esc(s: str) -> str:
        return s.replace("'", "''")

    for classe, items in DATA.items():
        _, systeme, niveau = META[classe]
        sys_id = annee_systeme.get(systeme, "CM-fr")
        for cat, article, qte, spec in items:
            nom_full = article + ((" — " + spec) if spec else "")
            nom_e = esc(nom_full)
            nom_norm_e = esc(normalize(article))
            niveau_e = esc(niveau)
            classe_e = esc(classe)
            lines_sql.append(
                f"INSERT INTO accessoires_populaires_par_classe "
                f"(pays, systeme_id, niveau, classe, nom, nom_normalise, "
                f"quantite_mediane, occurrences, derniere_vue) VALUES "
                f"('{pays}', '{sys_id}', '{niveau_e}', '{classe_e}', "
                f"'{nom_e}', '{nom_norm_e}', {qte}, 1, NOW()) "
                f"ON CONFLICT (pays, niveau, classe, nom_normalise) DO UPDATE SET "
                f"quantite_mediane = EXCLUDED.quantite_mediane, "
                f"occurrences = accessoires_populaires_par_classe.occurrences + 1, "
                f"derniere_vue = NOW();\n"
            )

    lines_sql.append(
        "\n-- Vérification :\n"
        "-- SELECT COUNT(*) AS total FROM accessoires_populaires_par_classe;\n"
        "-- SELECT classe, COUNT(*) FROM accessoires_populaires_par_classe\n"
        "--   GROUP BY classe ORDER BY classe;\n"
    )

    sql_path.write_text("".join(lines_sql), encoding="utf-8")

    # Copie aussi vers backend/migrations/ pour application auto au prochain
    # déploiement (sqlx::migrate!() embarque tous les .sql du dossier).
    migration_path = Path(__file__).parent.parent.parent / "backend" / "migrations" / "20260523_001_seed_accessoires_par_classe.sql"
    migration_path.write_text("".join(lines_sql), encoding="utf-8")

    print(f"   SQL seed -> {sql_path} ({total_lines} INSERTs)")
    print(f"   Migration -> {migration_path}")


if __name__ == "__main__":
    main()
