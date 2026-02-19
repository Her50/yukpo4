#!/usr/bin/env python3
"""
Script d'analyse approfondie d'un fichier Excel
Extrait : formules, liens entre feuilles, structure, dépendances
"""

import openpyxl
from openpyxl.formula.translate import Translator
import json
import re
from collections import defaultdict
from pathlib import Path
import sys

def extract_cell_references(formula):
    """Extrait les références de cellules d'une formule"""
    # Pattern pour références : Feuille!A1, A1, $A$1, etc.
    patterns = [
        r"([A-Za-z0-9_]+)!([$]?[A-Z]+[$]?[0-9]+)",  # Feuille!A1
        r"([$]?[A-Z]+[$]?[0-9]+)",  # A1 ou $A$1
        r"([A-Za-z0-9_]+)!([$]?[A-Z]+[$]?[0-9]+):([$]?[A-Z]+[$]?[0-9]+)",  # Feuille!A1:B2
    ]
    references = []
    for pattern in patterns:
        matches = re.findall(pattern, formula)
        references.extend(matches)
    return references

def analyze_excel_file(file_path):
    """Analyse complète d'un fichier Excel"""
    
    print(f"[ANALYSE] Fichier : {file_path}")
    print("=" * 80)
    
    try:
        wb = openpyxl.load_workbook(file_path, data_only=False)  # data_only=False pour garder les formules
    except Exception as e:
        print(f"[ERREUR] Ouverture : {e}")
        return None
    
    analysis = {
        "file_name": Path(file_path).name,
        "sheets": {},
        "cross_sheet_references": [],
        "formulas_summary": {},
        "hidden_sheets": [],
        "named_ranges": [],
        "structure": {}
    }
    
    # Liste toutes les feuilles (visibles et masquées)
    print(f"\n[FEUILLES] Detectees ({len(wb.sheetnames)} total)")
    print("-" * 80)
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        is_hidden = wb[ws.title].sheet_state == 'hidden'
        visibility = "[MASQUEE]" if is_hidden else "[VISIBLE]"
        print(f"  {visibility}: {sheet_name}")
        
        if is_hidden:
            analysis["hidden_sheets"].append(sheet_name)
    
    # Analyse détaillée de chaque feuille
    print(f"\n[ANALYSE] Detaillee par feuille")
    print("=" * 80)
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n[FEUILLE] {sheet_name}")
        print("-" * 80)
        
        sheet_data = {
            "name": sheet_name,
            "hidden": wb[sheet_name].sheet_state == 'hidden',
            "dimensions": f"{ws.max_row} lignes x {ws.max_column} colonnes",
            "formulas": [],
            "values": [],
            "cross_references": [],
            "data_regions": []
        }
        
        # Parcourir toutes les cellules
        formulas_in_sheet = []
        values_in_sheet = []
        cross_refs = []
        
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, 
                                min_col=1, max_col=ws.max_column):
            for cell in row:
                if cell.value is not None:
                    cell_info = {
                        "cell": cell.coordinate,
                        "value": str(cell.value)[:100]  # Limiter la longueur
                    }
                    
                    # Vérifier si c'est une formule
                    if cell.data_type == 'f':  # 'f' = formula
                        formula = cell.value
                        cell_info["formula"] = formula
                        cell_info["type"] = "formula"
                        formulas_in_sheet.append(cell_info)
                        
                        # Extraire les références
                        refs = extract_cell_references(formula)
                        if refs:
                            cell_info["references"] = refs
                        
                        # Détecter les références croisées (autres feuilles)
                        if '!' in formula:
                            for ref in refs:
                                if isinstance(ref, tuple) and len(ref) >= 2:
                                    other_sheet = ref[0]
                                    if other_sheet != sheet_name:
                                        cross_ref = {
                                            "from": f"{sheet_name}!{cell.coordinate}",
                                            "to": f"{other_sheet}!{ref[1] if len(ref) > 1 else ''}",
                                            "formula": formula
                                        }
                                        cross_refs.append(cross_ref)
                                        analysis["cross_sheet_references"].append(cross_ref)
                    else:
                        cell_info["type"] = "value"
                        values_in_sheet.append(cell_info)
        
        sheet_data["formulas"] = formulas_in_sheet
        sheet_data["values"] = values_in_sheet[:50]  # Limiter à 50 valeurs pour le JSON
        sheet_data["cross_references"] = cross_refs
        
        # Statistiques
        print(f"  [DIMENSIONS] {sheet_data['dimensions']}")
        print(f"  [FORMULES] {len(formulas_in_sheet)}")
        print(f"  [VALEURS] {len(values_in_sheet)}")
        print(f"  [REFERENCES CROISEES] {len(cross_refs)}")
        
        if formulas_in_sheet:
            print(f"\n  [EXEMPLES] Formules:")
            for i, f in enumerate(formulas_in_sheet[:5], 1):
                print(f"    {i}. {f['cell']}: {f.get('formula', '')[:80]}")
        
        if cross_refs:
            print(f"\n  [REFERENCES] Vers autres feuilles:")
            for ref in cross_refs[:5]:
                print(f"    {ref['from']} -> {ref['to']}")
        
        analysis["sheets"][sheet_name] = sheet_data
    
    # Analyse des plages nommées
    print(f"\n[PLAGES NOMEES]")
    print("-" * 80)
    if wb.defined_names:
        for name, definition in wb.defined_names.items():
            named_range = {
                "name": name,
                "definition": str(definition)
            }
            analysis["named_ranges"].append(named_range)
            print(f"  - {name}: {definition}")
    else:
        print("  Aucune plage nommee trouvee")
    
    # Résumé des formules par type
    print(f"\n[RESUME] Formules par type")
    print("-" * 80)
    formula_types = defaultdict(int)
    for sheet_name, sheet_data in analysis["sheets"].items():
        for formula_info in sheet_data["formulas"]:
            formula = formula_info.get("formula", "")
            # Détecter le type de formule
            if formula.startswith("="):
                func_match = re.match(r"=([A-Z]+)\(", formula)
                if func_match:
                    func_name = func_match.group(1)
                    formula_types[func_name] += 1
    
    for func_name, count in sorted(formula_types.items(), key=lambda x: -x[1]):
        print(f"  {func_name}: {count} utilisations")
    
    analysis["formulas_summary"] = dict(formula_types)
    
    # Structure générale
    analysis["structure"] = {
        "total_sheets": len(wb.sheetnames),
        "visible_sheets": len(wb.sheetnames) - len(analysis["hidden_sheets"]),
        "hidden_sheets": len(analysis["hidden_sheets"]),
        "total_formulas": sum(len(s["formulas"]) for s in analysis["sheets"].values()),
        "total_cross_references": len(analysis["cross_sheet_references"])
    }
    
    print(f"\n[RESUME GENERAL]")
    print("=" * 80)
    print(f"  Total feuilles: {analysis['structure']['total_sheets']}")
    print(f"  Feuilles visibles: {analysis['structure']['visible_sheets']}")
    print(f"  Feuilles masquees: {analysis['structure']['hidden_sheets']}")
    print(f"  Total formules: {analysis['structure']['total_formulas']}")
    print(f"  References croisees: {analysis['structure']['total_cross_references']}")
    
    return analysis

def save_analysis_report(analysis, output_file):
    """Sauvegarde le rapport d'analyse"""
    # Version JSON complète
    json_file = output_file.replace('.txt', '.json')
    with open(json_file, 'w', encoding='utf-8') as f:
        json.dump(analysis, f, indent=2, ensure_ascii=False)
    print(f"\n[SAUVEGARDE] Rapport JSON: {json_file}")
    
    # Version texte lisible
    with open(output_file, 'w', encoding='utf-8') as f:
        f.write("=" * 80 + "\n")
        f.write("RAPPORT D'ANALYSE EXCEL\n")
        f.write("=" * 80 + "\n\n")
        
        f.write(f"Fichier analysé: {analysis['file_name']}\n\n")
        
        f.write("STRUCTURE GÉNÉRALE\n")
        f.write("-" * 80 + "\n")
        for key, value in analysis['structure'].items():
            f.write(f"  {key}: {value}\n")
        f.write("\n")
        
        f.write("FEUILLES MASQUÉES\n")
        f.write("-" * 80 + "\n")
        if analysis['hidden_sheets']:
            for sheet in analysis['hidden_sheets']:
                f.write(f"  • {sheet}\n")
        else:
            f.write("  Aucune\n")
        f.write("\n")
        
        f.write("RÉFÉRENCES CROISÉES ENTRE FEUILLES\n")
        f.write("-" * 80 + "\n")
        for ref in analysis['cross_sheet_references'][:20]:  # Limiter à 20
            f.write(f"  {ref['from']} → {ref['to']}\n")
            f.write(f"    Formule: {ref['formula'][:100]}\n\n")
        
        f.write("\nFORMULES PAR FEUILLE\n")
        f.write("-" * 80 + "\n")
        for sheet_name, sheet_data in analysis['sheets'].items():
            f.write(f"\n{sheet_name}:\n")
            f.write(f"  Dimensions: {sheet_data['dimensions']}\n")
            f.write(f"  Formules: {len(sheet_data['formulas'])}\n")
            if sheet_data['formulas']:
                for formula in sheet_data['formulas'][:10]:  # Limiter à 10 par feuille
                    f.write(f"    {formula['cell']}: {formula.get('formula', '')[:100]}\n")
    
    print(f"[SAUVEGARDE] Rapport texte: {output_file}")

if __name__ == "__main__":
    # Chemin du fichier Excel
    excel_file = "gestion_microfinance2025_2027_test.xlsx"
    
    if not Path(excel_file).exists():
        print(f"[ERREUR] Fichier non trouve: {excel_file}")
        sys.exit(1)
    
    # Analyser le fichier
    analysis = analyze_excel_file(excel_file)
    
    if analysis:
        # Sauvegarder le rapport
        output_txt = "billing-support-info/excel_analysis_report.txt"
        save_analysis_report(analysis, output_txt)
        
        print(f"\n[SUCCES] Analyse terminee!")
        print(f"[RAPPORTS] Consultez: billing-support-info/")
