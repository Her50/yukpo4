#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script pour générer le fichier Excel des données financières Yukpo
Historique + Projections (2026-2030)
"""

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Installing openpyxl...")
    import subprocess
    import sys
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

def create_financial_excel():
    """Crée le fichier Excel complet avec toutes les données financières"""
    
    wb = Workbook()
    
    # Supprimer la feuille par défaut
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Styles
    header_fill = PatternFill(start_color="FF6600", end_color="FF6600", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    title_font = Font(bold=True, size=14, color="FF6600")
    subtitle_font = Font(bold=True, size=12)
    number_font = Font(size=10)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # ============================================
    # FEUILLE 1 : VUE D'ENSEMBLE
    # ============================================
    ws1 = wb.create_sheet("Vue d'ensemble", 0)
    ws1['A1'] = "YUKPO - Données Financières"
    ws1['A1'].font = Font(bold=True, size=16, color="FF6600")
    ws1['A2'] = "Seed Round : 1,1M EUR (≈1,2M USD, ≈720M FCFA)"
    ws1['A2'].font = subtitle_font
    
    row = 4
    ws1[f'A{row}'] = "INFORMATIONS GÉNÉRALES"
    ws1[f'A{row}'].font = title_font
    row += 2
    
    info_data = [
        ["Entreprise", "Yukpo"],
        ["Secteur", "Plateforme O2O Multi-Secteurs avec IA"],
        ["Marché cible", "Afrique Francophone (Cameroun, Sénégal, Côte d'Ivoire)"],
        ["Devise", "EUR (Euros)"],
        ["Taux de change", "1 EUR = 655 FCFA | 1 EUR = 1,09 USD"],
        ["Seed Round demandé", "1,1M EUR"],
        ["Objectif", "Acquisition utilisateurs + Validation marché"],
        ["Rentabilité prévue", "2028"],
    ]
    
    for info in info_data:
        ws1[f'A{row}'] = info[0]
        ws1[f'A{row}'].font = Font(bold=True)
        ws1[f'B{row}'] = info[1]
        row += 1
    
    row += 2
    ws1[f'A{row}'] = "STRUCTURE DU DOCUMENT"
    ws1[f'A{row}'].font = title_font
    row += 2
    
    structure = [
        ["Feuille 2", "Projections Financières (2026-2030)"],
        ["Feuille 3", "Détail des Charges"],
        ["Feuille 4", "Croissance Utilisateurs/Actifs"],
        ["Feuille 5", "Métriques Clés (CAC, LTV, Unit Economics)"],
        ["Feuille 6", "Données Historiques"],
    ]
    
    for struct in structure:
        ws1[f'A{row}'] = struct[0]
        ws1[f'A{row}'].font = Font(bold=True)
        ws1[f'B{row}'] = struct[1]
        row += 1
    
    # Ajuster largeur colonnes
    ws1.column_dimensions['A'].width = 25
    ws1.column_dimensions['B'].width = 50
    
    # ============================================
    # FEUILLE 2 : PROJECTIONS FINANCIÈRES
    # ============================================
    ws2 = wb.create_sheet("Projections Financières")
    
    headers = [
        "ANNÉE", "ACTIFS", "REVENUS PLACEMENT (K EUR)", "REVENUS AUTRES (K EUR)",
        "TOTAL REVENUS (K EUR)", "MARKETING (K EUR)", "CHARGES OPÉ. (K EUR)",
        "TOTAL CHARGES (K EUR)", "RÉSULTAT NET (K EUR)", "MARGE %"
    ]
    
    # En-têtes
    for col, header in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    
    # Données (toutes en K EUR)
    data = [
        ["2026 (6m)", 12950, 159, 53, 213, 337, 422, 759, -546, -257],
        ["2027", 40761, 1160, 217, 1378, 367, 1440, 1807, -433, -32],
        ["2028", 122283, 3500, 581, 4081, 459, 3600, 4059, 20, 0.5],
        ["2029", 305707, 8970, 1590, 10560, 734, 7220, 7954, 2600, 24.7],
        ["2030", 611415, 18700, 3600, 22300, 1100, 13420, 14520, 7780, 34.8],
    ]
    
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            cell.alignment = Alignment(horizontal='right' if col_idx > 1 else 'left', vertical='center')
            
            # Formatage des nombres
            if col_idx in [2, 3, 4, 5, 6, 7, 8, 9]:
                cell.number_format = '#,##0'
                if col_idx == 9 and value < 0:
                    cell.font = Font(size=10, color="DC2626")
                elif col_idx == 9 and value > 0:
                    cell.font = Font(size=10, color="059669")
            elif col_idx == 10:
                cell.number_format = '0.0"%"'
                if value < 0:
                    cell.font = Font(size=10, color="DC2626")
                elif value > 0:
                    cell.font = Font(size=10, color="059669")
            else:
                cell.font = Font(bold=True)
    
    # Ligne totale (optionnel pour 5 ans)
    total_row = 7
    ws2.cell(row=total_row, column=1, value="TOTAL 5 ANS").font = Font(bold=True)
    ws2.cell(row=total_row, column=4, value=38332).number_format = '#,##0'  # Total revenus
    ws2.cell(row=total_row, column=8, value=29099).number_format = '#,##0'  # Total charges
    ws2.cell(row=total_row, column=9, value=9221).number_format = '#,##0'  # Total résultat
    for col in range(1, 11):
        ws2.cell(row=total_row, column=col).border = border
        ws2.cell(row=total_row, column=col).fill = PatternFill(start_color="FFF4E6", end_color="FFF4E6", fill_type="solid")
    
    # Ajuster largeur colonnes
    for col in range(1, 11):
        ws2.column_dimensions[get_column_letter(col)].width = 18
    
    # ============================================
    # FEUILLE 3 : DÉTAIL DES CHARGES
    # ============================================
    ws3 = wb.create_sheet("Détail Charges")
    
    headers_charges = [
        "ANNÉE", "MARKETING (K EUR)", "PERSONNEL (K EUR)", "INFRASTRUCTURE (K EUR)",
        "OPÉRATIONNELLES (K EUR)", "TOTAL (K EUR)"
    ]
    
    for col, header in enumerate(headers_charges, 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    
    charges_data = [
        ["2026 (10m)", 337, 188, 76, 158, 759],
        ["2027", 367, 550, 229, 660, 1807],
        ["2028", 459, 1280, 642, 1678, 4059],
        ["2029", 734, 2750, 1378, 3092, 7954],
        ["2030", 1100, 4890, 2750, 5780, 14520],
    ]
    
    for row_idx, row_data in enumerate(charges_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws3.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            cell.alignment = Alignment(horizontal='right' if col_idx > 1 else 'left', vertical='center')
            cell.number_format = '#,##0' if col_idx > 1 else ''
            if col_idx == 1:
                cell.font = Font(bold=True)
    
    # Ajuster largeur colonnes
    for col in range(1, 7):
        ws3.column_dimensions[get_column_letter(col)].width = 20
    
    # ============================================
    # FEUILLE 4 : CROISSANCE UTILISATEURS/ACTIFS
    # ============================================
    ws4 = wb.create_sheet("Croissance Utilisateurs")
    
    headers_croissance = [
        "ANNÉE", "COMMERÇANTS", "PRESTATAIRES", "TOTAL ACTIFS", "UTILISATEURS"
    ]
    
    for col, header in enumerate(headers_croissance, 1):
        cell = ws4.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    
    croissance_data = [
        ["2026 (6m)", 10000, 2950, 12950, 500000],
        ["2027", 28532, 12228, 40761, 1500000],
        ["2028", 85598, 36684, 122283, 4000000],
        ["2029", 213994, 91712, 305707, 12000000],
        ["2030", 427990, 183424, 611415, 25000000],
    ]
    
    for row_idx, row_data in enumerate(croissance_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws4.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            cell.alignment = Alignment(horizontal='right' if col_idx > 1 else 'left', vertical='center')
            cell.number_format = '#,##0' if col_idx > 1 else ''
            if col_idx == 1:
                cell.font = Font(bold=True)
    
    # Ajuster largeur colonnes
    for col in range(1, 6):
        ws4.column_dimensions[get_column_letter(col)].width = 20
    
    # ============================================
    # FEUILLE 5 : MÉTRIQUES CLÉS
    # ============================================
    ws5 = wb.create_sheet("Métriques Clés")
    
    ws5['A1'] = "MÉTRIQUES CLÉS DE VALIDATION SEED"
    ws5['A1'].font = title_font
    ws5.merge_cells('A1:B1')
    
    row = 3
    metrics = [
        ["LTV (Lifetime Value)", "73 EUR (48K FCFA) - Rétention 24 mois × 3,05 EUR/mois"],
        ["CAC Commerçants", "23-39 EUR (15K-25K FCFA)"],
        ["CAC Utilisateurs", "0,76-1,53 EUR (500-1K FCFA)"],
        ["LTV/CAC Ratio", "2,4x (objectif Seed : >3x avec optimisation)"],
        ["Payback Period", "10 mois (CAC récupéré)"],
        ["Unit Economics", "Marge unitaire : 3,05 EUR/mois (2 000 FCFA/mois)"],
        ["Taux d'adoption 2026", "0,35% marché camerounais"],
        ["Taux d'adoption 2027", "0,63% marché camerounais"],
        ["Taux d'adoption 2030", "9,45% marché camerounais"],
        ["Rétention cible", "80%+ après 3 mois"],
    ]
    
    for metric in metrics:
        ws5[f'A{row}'] = metric[0]
        ws5[f'A{row}'].font = Font(bold=True)
        ws5[f'B{row}'] = metric[1]
        row += 1
    
    row += 2
    ws5[f'A{row}'] = "MILESTONES SEED"
    ws5[f'A{row}'].font = title_font
    ws5.merge_cells(f'A{row}:B{row}')
    row += 2
    
    milestones = [
        ["M+6 (Août 2026)", "12,950 actifs | 213K EUR revenus"],
        ["M+12 (Février 2027)", "40,761 actifs | 1,38M EUR revenus"],
        ["M+18 (Août 2027)", "Product-Market Fit validé"],
        ["M+24 (Février 2028)", "Rentabilité opérationnelle atteinte"],
    ]
    
    for milestone in milestones:
        ws5[f'A{row}'] = milestone[0]
        ws5[f'A{row}'].font = Font(bold=True)
        ws5[f'B{row}'] = milestone[1]
        row += 1
    
    ws5.column_dimensions['A'].width = 30
    ws5.column_dimensions['B'].width = 50
    
    # ============================================
    # FEUILLE 6 : DONNÉES HISTORIQUES
    # ============================================
    ws6 = wb.create_sheet("Historique")
    
    ws6['A1'] = "DONNÉES FINANCIÈRES HISTORIQUES"
    ws6['A1'].font = title_font
    ws6.merge_cells('A1:D1')
    
    row = 3
    ws6[f'A{row}'] = "NOTE IMPORTANTE"
    ws6[f'A{row}'].font = Font(bold=True, color="DC2626")
    ws6[f'B{row}'] = "Yukpo est actuellement en phase pré-revenus (produit opérationnel, pas encore de traction commerciale significative)"
    ws6.merge_cells(f'B{row}:D{row}')
    row += 2
    
    ws6[f'A{row}'] = "PÉRIODE"
    ws6[f'A{row}'].font = header_font
    ws6[f'A{row}'].fill = header_fill
    ws6[f'A{row}'].border = border
    ws6[f'B{row}'] = "REVENUS (EUR)"
    ws6[f'B{row}'].font = header_font
    ws6[f'B{row}'].fill = header_fill
    ws6[f'B{row}'].border = border
    ws6[f'C{row}'] = "CHARGES (EUR)"
    ws6[f'C{row}'].font = header_font
    ws6[f'C{row}'].fill = header_fill
    ws6[f'C{row}'].border = border
    ws6[f'D{row}'] = "RÉSULTAT NET (EUR)"
    ws6[f'D{row}'].font = header_font
    ws6[f'D{row}'].fill = header_fill
    ws6[f'D{row}'].border = border
    row += 1
    
    historique_data = [
        ["Mars 2025 - Déc 2025", 0, 500, -500],
        ["Jan 2026 - Fév 2026", 0, 1000, -1000],
    ]
    
    for hist_data in historique_data:
        ws6[f'A{row}'] = hist_data[0]
        ws6[f'B{row}'] = hist_data[1]
        ws6[f'C{row}'] = hist_data[2]
        ws6[f'D{row}'] = hist_data[3]
        for col in range(1, 5):
            cell = ws6.cell(row=row, column=col)
            cell.border = border
            cell.alignment = Alignment(horizontal='right' if col > 1 else 'left', vertical='center')
            if col > 1:
                cell.number_format = '#,##0'
                if col == 4 and hist_data[3] < 0:
                    cell.font = Font(size=10, color="DC2626")
        row += 1
    
    row += 1
    ws6[f'A{row}'] = "DÉTAIL DES CHARGES HISTORIQUES"
    ws6[f'A{row}'].font = subtitle_font
    ws6.merge_cells(f'A{row}:D{row}')
    row += 2
    
    ws6[f'A{row}'] = "Poste"
    ws6[f'A{row}'].font = Font(bold=True)
    ws6[f'B{row}'] = "Montant (EUR)"
    ws6[f'B{row}'].font = Font(bold=True)
    row += 1
    
    charges_hist = [
        ["Infrastructure Cloud (Render, AWS)", "1 500 EUR"],
        ["Services IA (OpenAI, Mistral)", "0 EUR (crédits gratuits utilisés)"],
        ["Domaine & SSL", "50 EUR"],
        ["TOTAL", "1 550 EUR"],
    ]
    
    for charge in charges_hist:
        ws6[f'A{row}'] = charge[0]
        ws6[f'B{row}'] = charge[1]
        if charge[0] == "TOTAL":
            ws6[f'A{row}'].font = Font(bold=True)
            ws6[f'B{row}'].font = Font(bold=True)
        row += 1
    
    ws6.column_dimensions['A'].width = 40
    ws6.column_dimensions['B'].width = 20
    ws6.column_dimensions['C'].width = 20
    ws6.column_dimensions['D'].width = 20
    
    # ============================================
    # ENREGISTREMENT
    # ============================================
    import os
    filename = "YUKPO_Donnees_Financieres_Orange.xlsx"
    # Chemin relatif au dossier courant (ORANGE)
    filepath = os.path.join(os.path.dirname(__file__), filename)
    wb.save(filepath)
    print(f"Fichier Excel cree : {filepath}")
    return filepath

if __name__ == "__main__":
    try:
        filepath = create_financial_excel()
        print(f"\nFichier Excel genere avec succes !")
        print(f"Emplacement : {filepath}")
        print(f"\nContenu :")
        print(f"  - Feuille 1 : Vue d'ensemble")
        print(f"  - Feuille 2 : Projections Financieres (2026-2030)")
        print(f"  - Feuille 3 : Detail des Charges")
        print(f"  - Feuille 4 : Croissance Utilisateurs/Actifs")
        print(f"  - Feuille 5 : Metriques Cles")
        print(f"  - Feuille 6 : Donnees Historiques")
    except Exception as e:
        print(f"Erreur : {e}")
        import traceback
        traceback.print_exc()

